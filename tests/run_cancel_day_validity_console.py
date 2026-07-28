"""Validate a read-only cancel-day prototype without editing source workbooks."""

from __future__ import annotations

import argparse
import hashlib
import importlib
import json
import shutil
import sys
import tempfile
import time
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from tools.cancel_day import cancel_day
from tools.get_schedule import FAKE_DATA_DIR


HEADERS = [
    "Session ID",
    "Session Type",
    "Course ID",
    "Course Name",
    "Student Groups",
    "Staff",
    "Room",
    "Room Type",
    "Room Capacity",
    "Expected Students",
    "Day",
    "Week",
    "Period",
    "Start",
    "End",
    "Status",
]


def _arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate and validate a read-only cancellation prototype."
    )
    parser.add_argument("--day", default="Monday")
    parser.add_argument("--week", type=int, default=1)
    parser.add_argument("--following-weeks", type=int, choices=(1, 2), default=2)
    parser.add_argument("--max-issues", type=int, default=1000)
    return parser.parse_args()


def _hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _cell(value: Any) -> Any:
    if isinstance(value, (list, tuple, set)):
        return "; ".join(str(item) for item in value)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return value


def _candidate_row(row: dict[str, Any]) -> list[Any]:
    unique_id = "COMP::{session_id}::W{week}::{day}::{period}".format(
        session_id=row.get("session_id"),
        week=row.get("academic_week"),
        day=row.get("day"),
        period=row.get("period_id") or row.get("period"),
    )
    values = [
        unique_id,
        row.get("session_type"),
        row.get("course_id"),
        row.get("course_name"),
        row.get("student_groups"),
        row.get("staff"),
        row.get("room"),
        row.get("room_type"),
        row.get("room_capacity"),
        row.get("expected_students"),
        row.get("day"),
        row.get("academic_week"),
        row.get("period_id") or row.get("period"),
        row.get("start"),
        row.get("end"),
        "Reserved",
    ]
    return [_cell(value) for value in values]


def _add_table(workbook: Workbook, title: str, table_name: str, rows: list[list[Any]]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(HEADERS)
    for row in rows:
        sheet.append(row)
    table = Table(displayName=table_name, ref=f"A1:P{len(rows) + 1}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def _doctor_names(path: Path) -> set[str]:
    workbook = load_workbook(path, read_only=False, data_only=True)
    try:
        sheet = workbook["Doctor Directory"]
        return {
            str(sheet.cell(row=row_number, column=2).value).strip().casefold()
            for row_number in range(5, sheet.max_row + 1)
            if sheet.cell(row=row_number, column=2).value not in (None, "")
        }
    finally:
        workbook.close()


def _staff_members(value: Any) -> list[str]:
    if isinstance(value, (list, tuple, set)):
        return [str(item).strip() for item in value if str(item).strip()]
    return [
        item.strip()
        for item in str(value or "").replace(",", ";").split(";")
        if item.strip()
    ]


def _write_candidate(
    path: Path,
    prototype_rows: list[dict[str, Any]],
    known_doctors: set[str],
) -> None:
    rows = [_candidate_row(row) for row in prototype_rows]
    doctor_rows: list[list[Any]] = []
    for prototype_row, candidate_row in zip(prototype_rows, rows):
        doctors = [
            member
            for member in _staff_members(prototype_row.get("staff"))
            if member.casefold() in known_doctors
        ]
        if doctors:
            doctor_row = list(candidate_row)
            doctor_row[5] = "; ".join(doctors)
            doctor_rows.append(doctor_row)
    workbook = Workbook()
    workbook.remove(workbook.active)
    _add_table(workbook, "Proposed Sessions", "ProposedSessions", rows)
    _add_table(workbook, "Proposed Room Reservations", "ProposedRooms", rows)
    _add_table(
        workbook,
        "Proposed Doctor Schedule",
        "ProposedStaff",
        doctor_rows,
    )
    workbook.save(path)
    workbook.close()


def _invoke_prototype(arguments: dict[str, Any]) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    first = json.loads(cancel_day.invoke(arguments))
    timetable = first.get("prototype_timetable") or {}
    rows = list(timetable.get("sessions") or [])
    pagination = timetable.get("pagination") or {}
    while pagination.get("has_more"):
        page_arguments = dict(arguments)
        page_arguments["result_offset"] = pagination["next_result_offset"]
        page = json.loads(cancel_day.invoke(page_arguments))
        page_timetable = page.get("prototype_timetable") or {}
        rows.extend(page_timetable.get("sessions") or [])
        pagination = page_timetable.get("pagination") or {}
    return first, rows


def _positive_delta(after: dict[str, int], before: dict[str, int]) -> dict[str, int]:
    return {
        key: value - before.get(key, 0)
        for key, value in sorted(after.items())
        if value - before.get(key, 0) > 0
    }


def main() -> int:
    arguments = _arguments()
    source_names = [
        "05_General_Schedule.xlsx",
        "Doctor Schedule Calendar.xlsx",
        "01_Room_Schedule.xlsx",
        "06_Exam_Schedule.xlsx",
    ]
    source_paths = [FAKE_DATA_DIR / name for name in source_names]
    hashes_before = {path.name: _hash(path) for path in source_paths}
    started = time.perf_counter()
    prototype, prototype_rows = _invoke_prototype(
        {
            "day": arguments.day,
            "academic_week": arguments.week,
            "reason": "Confirmed full-campus closure validity test.",
            "cancellation_approved": True,
            "maximum_following_weeks": arguments.following_weeks,
            "result_offset": 0,
            "result_limit": 100,
        }
    )

    validity_module = importlib.import_module("tools.check_validity")
    with tempfile.TemporaryDirectory(prefix="cancel-day-validity-") as temp_name:
        # check_validity deliberately rejects paths outside its configured
        # data root, so use the same resolved Windows path on both sides.
        temp_path = Path(temp_name).resolve()
        for source in source_paths:
            shutil.copy2(source, temp_path / source.name)
        candidate_name = "Monday_Cancellation_Prototype.xlsx"
        _write_candidate(
            temp_path / candidate_name,
            prototype_rows,
            _doctor_names(FAKE_DATA_DIR / "Doctor Schedule Calendar.xlsx"),
        )
        validity_module.FAKE_DATA_DIR = temp_path

        baseline = json.loads(
            validity_module.check_validity.invoke(
                {"schedule_files": source_names, "max_issues": arguments.max_issues}
            )
        )
        candidate_mappings = {
            f"{candidate_name}::Proposed Sessions": {"_role": "sessions"},
            f"{candidate_name}::Proposed Room Reservations": {
                "_role": "room_availability"
            },
            f"{candidate_name}::Proposed Doctor Schedule": {
                "_role": "doctor_sessions"
            },
        }
        combined = json.loads(
            validity_module.check_validity.invoke(
                {
                    "schedule_files": [*source_names, candidate_name],
                    "column_mappings": candidate_mappings,
                    "max_issues": arguments.max_issues,
                }
            )
        )

    hashes_after = {path.name: _hash(path) for path in source_paths}
    baseline_summary = baseline.get("summary") or {}
    combined_summary = combined.get("summary") or {}
    code_delta = _positive_delta(
        combined_summary.get("issues_by_code") or {},
        baseline_summary.get("issues_by_code") or {},
    )
    severity_delta = _positive_delta(
        combined_summary.get("issues_by_severity") or {},
        baseline_summary.get("issues_by_severity") or {},
    )
    result = {
        "prototype_id": prototype.get("prototype_id"),
        "prototype_status": prototype.get("status"),
        "affected_session_count": (prototype.get("cancelled_scope") or {}).get(
            "affected_session_count"
        ),
        "proposed_session_count": len(prototype_rows),
        "unassigned_session_count": prototype.get("unassigned_session_count"),
        "baseline_validation_status": baseline.get("validation_status"),
        "combined_validation_status": combined.get("validation_status"),
        "baseline_total_issues": baseline_summary.get("total_issues"),
        "combined_total_issues": combined_summary.get("total_issues"),
        "prototype_added_issues_by_severity": severity_delta,
        "prototype_added_issues_by_code": code_delta,
        "prototype_passed_supported_conflict_checks": not code_delta,
        "combined_validation_complete": combined.get("validation_complete"),
        "combined_incomplete_reasons": combined.get("incomplete_reasons"),
        "baseline_source_errors": baseline.get("source_errors"),
        "combined_source_errors": combined.get("source_errors"),
        "baseline_extraction_errors": baseline.get("extraction_errors"),
        "combined_extraction_errors": combined.get("extraction_errors"),
        "source_files_modified": hashes_before != hashes_after,
        "elapsed_seconds": round(time.perf_counter() - started, 2),
    }
    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
    return 0 if not result["source_files_modified"] else 3


if __name__ == "__main__":
    raise SystemExit(main())
