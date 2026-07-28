"""Comprehensive contract tests for every implemented scheduler tool.

The two explicit TODO skeletons (``cancel_day`` and ``approve_repair``) are not
included here because they do not yet have behavior to verify.
"""

from __future__ import annotations

import importlib
import json
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table


get_schedule_module = importlib.import_module("tools.get_schedule")
find_module = importlib.import_module("tools.find_affected_sessions")
priority_module = importlib.import_module("tools.check_priority")
staff_module = importlib.import_module("tools.check_lecturer_or_ta_availability")
room_module = importlib.import_module("tools.check_room_availability")
compare_module = importlib.import_module("tools.compare_schedule_versions")
validity_module = importlib.import_module("tools.check_validity")


class FakeTool:
    def __init__(self, responder):
        self.responder = responder
        self.calls: list[dict] = []

    def invoke(self, arguments: dict):
        self.calls.append(arguments)
        result = self.responder(arguments) if callable(self.responder) else self.responder
        return result if isinstance(result, str) else json.dumps(result)


def tool_result(tool, **arguments) -> dict:
    return json.loads(tool.invoke(arguments))


def extraction_payload(rows: list[dict], *, sheet: str = "Semester Timetable") -> dict:
    return {
        "status": "ok",
        "extraction": {
            "matching_rows_found": len(rows),
            "matching_rows_returned": len(rows),
            "selected_sheets": [sheet],
            "has_more": False,
            "next_row_offset": None,
            "sheets": [
                {
                    "name": sheet,
                    "tables": [
                        {
                            "name": "Schedule",
                            "rows": [
                                {"excel_row": index + 2, "values": row}
                                for index, row in enumerate(rows)
                            ],
                        }
                    ],
                }
            ],
        },
        "limits": {"truncated": False, "completeness": "complete"},
    }


def affected_record(key: str, session_type: str) -> dict:
    return {
        "affected_session_key": key,
        "session_id": key,
        "course_id": f"C-{key}",
        "course_name": f"Course {key}",
        "session_type": session_type,
        "day": "Monday",
        "start": "08:30",
        "end": "10:00",
        "week": "Weeks 1-12",
        "room": "R1",
        "expected_students": 30,
        "student_groups": "G1",
        "instructor_id": "D1",
    }


class TemporaryWorkbookTestCase(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_context = tempfile.TemporaryDirectory()
        self.data_dir = Path(self.temp_context.name).resolve()

    def tearDown(self) -> None:
        self.temp_context.cleanup()

    def save_workbook(self, name: str, sheets: dict[str, list[list[object]]]) -> Path:
        workbook = Workbook()
        first = True
        for table_number, (sheet_name, rows) in enumerate(sheets.items(), start=1):
            sheet = workbook.active if first else workbook.create_sheet()
            first = False
            sheet.title = sheet_name
            for row in rows:
                sheet.append(row)
            if len(rows) >= 2 and rows[0]:
                reference = f"A1:{get_column_letter(len(rows[0]))}{len(rows)}"
                sheet.add_table(Table(displayName=f"ScheduleTable{table_number}", ref=reference))
        path = self.data_dir / name
        workbook.save(path)
        return path


class GetScheduleTests(TemporaryWorkbookTestCase):
    def setUp(self) -> None:
        super().setUp()
        self.save_workbook(
            "schedule.xlsx",
            {
                "Semester Timetable": [
                    ["Session ID", "Day", "Session Type", "Room"],
                    ["S1", "Monday", "Lecture", "R1"],
                    ["S2", "Monday", "Tutorial", "R2"],
                    ["S3", "Tuesday", "Lecture", "R3"],
                ]
            },
        )

    def test_excel_filters_and_pagination_are_complete(self) -> None:
        with patch.object(get_schedule_module, "FAKE_DATA_DIR", self.data_dir):
            first = tool_result(
                get_schedule_module.get_schedule,
                uploaded_file_path="schedule.xlsx",
                sheet_name="Semester Timetable",
                filters={"day": "Monday"},
                max_rows=1,
            )
            second = tool_result(
                get_schedule_module.get_schedule,
                uploaded_file_path="schedule.xlsx",
                sheet_name="Semester Timetable",
                filters={"day": "Monday"},
                row_offset=first["extraction"]["next_row_offset"],
                max_rows=1,
            )

        self.assertEqual(first["status"], "ok")
        self.assertEqual(first["extraction"]["matching_rows_found"], 2)
        self.assertTrue(first["extraction"]["has_more"])
        self.assertFalse(second["extraction"]["has_more"])

    def test_rejects_file_outside_permitted_folder(self) -> None:
        with patch.object(get_schedule_module, "FAKE_DATA_DIR", self.data_dir):
            result = tool_result(
                get_schedule_module.get_schedule,
                uploaded_file_path="../outside.xlsx",
            )
        self.assertEqual(result["status"], "error")
        self.assertEqual(result["error"]["code"], "path_outside_fake_data")


class FindAffectedSessionsTests(unittest.TestCase):
    def test_filters_week_and_inactive_sessions(self) -> None:
        rows = [
            {
                "Session ID": "S1",
                "Session Type": "Lecture",
                "Day": "Monday",
                "Start": "08:30",
                "End": "10:00",
                "Weeks": "Weeks 1-12",
                "Room": "R1",
                "Expected Students": 30,
                "Student Groups": "G1",
                "Instructor ID": "D1",
                "Status": "Active",
            },
            {
                "Session ID": "S2",
                "Session Type": "Tutorial",
                "Day": "Monday",
                "Start": "10:15",
                "End": "11:45",
                "Weeks": "Week 2",
                "Room": "R2",
                "Student Groups": "G2",
                "Instructor ID": "T1",
                "Status": "Active",
            },
            {
                "Session ID": "S3",
                "Session Type": "Lecture",
                "Day": "Monday",
                "Start": "12:00",
                "End": "13:30",
                "Weeks": "Weeks 1-12",
                "Room": "R3",
                "Student Groups": "G3",
                "Instructor ID": "D3",
                "Status": "Cancelled",
            },
        ]
        fake = FakeTool(extraction_payload(rows))
        with patch.object(find_module, "get_schedule", fake):
            result = tool_result(
                find_module.find_affected_sessions,
                uploaded_file_path="general.xlsx",
                affected_day_or_date="Monday",
                academic_week=1,
            )

        self.assertEqual(result["status"], "success")
        self.assertTrue(result["complete"])
        self.assertEqual(result["affected_session_count"], 1)
        self.assertEqual(result["affected_sessions"][0]["affected_session_key"], "S1")
        self.assertEqual(result["excluded_counts"]["already_inactive"], 1)

    def test_invalid_week_is_rejected_without_dependency_call(self) -> None:
        fake = FakeTool(extraction_payload([]))
        with patch.object(find_module, "get_schedule", fake):
            result = tool_result(
                find_module.find_affected_sessions,
                uploaded_file_path="general.xlsx",
                affected_day_or_date="Monday",
                academic_week=0,
            )
        self.assertEqual(result["status"], "invalid_request")
        self.assertEqual(fake.calls, [])


class CheckPriorityTests(unittest.TestCase):
    def _payload(self, sessions: list[dict]) -> dict:
        return {
            "status": "success",
            "complete": True,
            "affected_session_count": len(sessions),
            "affected_sessions": sessions,
            "result_pagination": {"has_more": False, "next_result_offset": None},
        }

    def test_strict_session_type_hierarchy(self) -> None:
        sessions = [
            affected_record("T", "Tutorial"),
            affected_record("L", "Laboratory"),
            affected_record("C", "Lecture"),
            affected_record("E", "Exam"),
        ]
        with patch.object(
            priority_module, "find_affected_sessions", FakeTool(self._payload(sessions))
        ):
            result = tool_result(
                priority_module.check_priority,
                uploaded_file_path="general.xlsx",
                affected_day_or_date="Monday",
                academic_week=1,
            )
        self.assertEqual(result["status"], "success")
        self.assertTrue(result["ranking_complete"])
        self.assertEqual(result["global_repair_order"], ["E", "C", "L", "T"])

    def test_unknown_type_requires_classification(self) -> None:
        with patch.object(
            priority_module,
            "find_affected_sessions",
            FakeTool(self._payload([affected_record("X", "Seminar")])),
        ):
            result = tool_result(
                priority_module.check_priority,
                uploaded_file_path="general.xlsx",
                affected_day_or_date="Monday",
                academic_week=1,
            )
        self.assertEqual(result["status"], "information_required")
        self.assertFalse(result["ranking_complete"])
        self.assertEqual(len(result["unclassified_sessions"]), 1)


class StaffAvailabilityTests(unittest.TestCase):
    def _staff_payload(self) -> dict:
        return extraction_payload(
            [
                {
                    "Staff ID": "D1",
                    "Day": "Monday",
                    "Start": "09:00",
                    "End": "10:00",
                    "Weeks": "Weeks 1-12",
                    "Course ID": "C1",
                    "Room": "R1",
                }
            ],
            sheet="D1",
        )

    def test_overlap_is_unavailable_but_adjacent_period_is_available(self) -> None:
        with patch.object(staff_module, "get_schedule", FakeTool(self._staff_payload())):
            overlap = tool_result(
                staff_module.check_lecturer_or_ta_availability,
                uploaded_file_path="staff.xlsx",
                staff_ids=["D1"],
                proposed_day="Monday",
                proposed_start="09:30",
                proposed_end="10:30",
                academic_week=1,
            )
        with patch.object(staff_module, "get_schedule", FakeTool(self._staff_payload())):
            adjacent = tool_result(
                staff_module.check_lecturer_or_ta_availability,
                uploaded_file_path="staff.xlsx",
                staff_ids=["D1"],
                proposed_day="Monday",
                proposed_start="10:00",
                proposed_end="11:00",
                academic_week=1,
            )
        self.assertFalse(overlap["all_available"])
        self.assertTrue(adjacent["all_available"])

    def test_unknown_staff_is_never_treated_as_available(self) -> None:
        empty = extraction_payload([], sheet="Directory")
        with patch.object(staff_module, "get_schedule", FakeTool(empty)):
            result = tool_result(
                staff_module.check_lecturer_or_ta_availability,
                uploaded_file_path="staff.xlsx",
                staff_ids=["UNKNOWN"],
                proposed_day="Monday",
                proposed_start="10:00",
                proposed_end="11:00",
                academic_week=1,
            )
        self.assertEqual(result["status"], "information_required")
        self.assertIsNone(result["all_available"])
        self.assertEqual(result["unknown_staff_ids"], ["UNKNOWN"])


class RoomAvailabilityTests(TemporaryWorkbookTestCase):
    def setUp(self) -> None:
        super().setUp()
        self.save_workbook(
            "rooms.xlsx",
            {
                "Inventory": [
                    ["Room", "Capacity", "Room Type", "Features"],
                    ["R1", 30, "Lecture Hall", "Projector"],
                    ["R2", 60, "Lecture Hall", "Projector, Accessible"],
                ],
                "Bookings": [
                    ["Room", "Day", "Start", "End", "Status"],
                    ["R1", "Monday", "09:00", "10:00", "Booked"],
                    ["R2", "Tuesday", "09:00", "10:00", "Booked"],
                ],
            },
        )

    def test_filters_conflicts_capacity_features_and_type(self) -> None:
        with patch.object(room_module, "FAKE_DATA_DIR", self.data_dir):
            result = tool_result(
                room_module.check_room_availability,
                uploaded_file_path="rooms.xlsx",
                requested_day_or_date="Monday",
                requested_start="09:00",
                requested_end="10:00",
                minimum_capacity=40,
                required_features=["Accessible"],
                room_types=["Lecture Hall"],
            )
        self.assertEqual(result["status"], "success")
        self.assertEqual([room["room"] for room in result["available_rooms"]], ["R2"])
        self.assertEqual(result["excluded_counts"]["time_conflict"], 1)

    def test_invalid_time_range_requests_information(self) -> None:
        with patch.object(room_module, "FAKE_DATA_DIR", self.data_dir):
            result = tool_result(
                room_module.check_room_availability,
                uploaded_file_path="rooms.xlsx",
                requested_day_or_date="Monday",
                requested_start="11:00",
                requested_end="10:00",
            )
        self.assertEqual(result["status"], "information_required")
        self.assertEqual(result["available_rooms"], [])


class CompareScheduleVersionsTests(unittest.TestCase):
    def test_reports_added_removed_modified_and_unchanged_sessions(self) -> None:
        original = [
            {"Session ID": "S1", "Day": "Monday", "Start": "09:00", "End": "10:00", "Room": "R1"},
            {"Session ID": "S2", "Day": "Tuesday", "Start": "09:00", "End": "10:00", "Room": "R2"},
            {"Session ID": "S3", "Day": "Wednesday", "Start": "09:00", "End": "10:00", "Room": "R3"},
        ]
        repaired = [
            {"Session ID": "S1", "Day": "Monday", "Start": "11:00", "End": "12:00", "Room": "R4"},
            {"Session ID": "S3", "Day": "Wednesday", "Start": "09:00", "End": "10:00", "Room": "R3"},
            {"Session ID": "S4", "Day": "Thursday", "Start": "09:00", "End": "10:00", "Room": "R4"},
        ]

        def retrieve(file_path: str, _sheet_names: list[str]):
            rows = original if file_path == "original.xlsx" else repaired
            return rows, [{"complete": True, "row_count": len(rows)}]

        with patch.object(compare_module, "_retrieve_version", side_effect=retrieve):
            result = tool_result(
                compare_module.compare_schedule_versions,
                original_file_path="original.xlsx",
                repaired_file_path="repaired.xlsx",
                sheet_names=["Schedule"],
                include_unchanged=True,
            )
        self.assertEqual(result["status"], "success")
        self.assertEqual(
            result["change_totals"],
            {"added": 1, "removed": 1, "modified": 1, "unchanged": 1, "changed": 3, "detail_records_available": 4},
        )
        statuses = {item["session_key"]: item["change_status"] for item in result["changes"]}
        self.assertEqual(statuses, {"S2": "removed", "S1": "modified", "S4": "added", "S3": "unchanged"})
        self.assertFalse(result["validation"]["repair_validated"])

    def test_requires_a_sheet_name(self) -> None:
        result = tool_result(
            compare_module.compare_schedule_versions,
            original_file_path="original.xlsx",
            repaired_file_path="repaired.xlsx",
            sheet_names=[],
        )
        self.assertEqual(result["status"], "invalid_request")


class CheckValidityTests(TemporaryWorkbookTestCase):
    def test_detects_room_staff_and_student_conflicts(self) -> None:
        self.save_workbook(
            "conflicts.xlsx",
            {
                "Schedule": [
                    ["Session ID", "Course ID", "Session Type", "Day", "Start", "End", "Weeks", "Room", "Expected Students", "Student Groups", "Instructor"],
                    ["S1", "C1", "Lecture", "Monday", "09:00", "10:00", "1", "R1", 30, "G1", "D1"],
                    ["S2", "C2", "Lecture", "Monday", "09:30", "10:30", "1", "R1", 25, "G1", "D1"],
                ]
            },
        )
        with patch.object(validity_module, "FAKE_DATA_DIR", self.data_dir):
            result = tool_result(
                validity_module.check_validity,
                schedule_files=["conflicts.xlsx"],
            )
        self.assertEqual(result["status"], "ok")
        self.assertEqual(result["validation_status"], "invalid")
        codes = result["summary"]["issues_by_code"]
        self.assertGreaterEqual(codes.get("ROOM_DOUBLE_BOOKING", 0), 1)
        self.assertGreaterEqual(codes.get("INSTRUCTOR_DOUBLE_BOOKING", 0), 1)
        self.assertGreaterEqual(codes.get("STUDENT_GROUP_DOUBLE_BOOKING", 0), 1)

    def test_no_files_and_invalid_rules_fail_safely(self) -> None:
        with patch.object(validity_module, "FAKE_DATA_DIR", self.data_dir):
            no_files = tool_result(validity_module.check_validity, schedule_files=[])
        with patch.object(validity_module, "FAKE_DATA_DIR", self.data_dir):
            bad_rules = tool_result(
                validity_module.check_validity,
                schedule_files=["missing.xlsx"],
                rules={"teaching_weeks": []},
            )
        self.assertEqual(no_files["status"], "error")
        self.assertEqual(no_files["error"]["code"], "no_schedule_files")
        self.assertEqual(bad_rules["error"]["code"], "invalid_rules")


if __name__ == "__main__":
    unittest.main()
