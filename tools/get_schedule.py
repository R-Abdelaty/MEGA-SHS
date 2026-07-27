"""Read-only extraction tool for schedule files stored in ``fake data``."""

from __future__ import annotations

import json
import re
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, cast

from langchain.tools import tool
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from pypdf import PdfReader


PROJECT_ROOT = Path(__file__).resolve().parent.parent
FAKE_DATA_DIR = (PROJECT_ROOT / "fake data").resolve()
SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".pdf"}
DEFAULT_MAX_ROWS = 100
MAX_ROW_LIMIT = 500
DEFAULT_MAX_PAGES = 10
MAX_PAGE_LIMIT = 50
DEFAULT_MAX_CHARS = 40_000
MAX_CHAR_LIMIT = 120_000
MAX_CELL_CHARS = 2_000
QUERY_STOP_WORDS = {
    "all",
    "at",
    "for",
    "from",
    "get",
    "in",
    "of",
    "on",
    "show",
    "the",
}
QUERY_SYNONYMS = {
    "tut": "tutorial",
    "tuts": "tutorial",
    "tutorials": "tutorial",
    "lectures": "lecture",
    "labs": "lab",
}
FILTER_HEADER_PREFERENCES: dict[str, list[str]] = {
    "student groups": [
        "Cohort Group(s)",
        "Student Group(s)",
        "Tutorial Group(s)",
        "Group ID",
        "Section",
    ],
    "session type": [
        "Session Type",
        "Activity Type",
        "Assessment Type",
        "Exam Type",
    ],
    "day": ["Day", "Weekday", "Day of Week"],
    "major": ["Major Code(s)", "Major(s)", "Major", "Program"],
    "course": ["Course ID", "Course Code", "Course Name", "Subject", "Module"],
    "room": ["Room", "Room ID", "Venue", "Location", "Hall", "Lab"],
    "instructor": ["Instructor", "Lecturer", "Doctor", "Staff", "Teacher", "TA"],
    "period": ["Period ID", "Period", "Slot", "Time Slot"],
    "week": ["Week", "Weeks", "Teaching Week", "Semester Week"],
}
FILTER_KEY_ALIASES = {
    "activity": "session type",
    "activity type": "session type",
    "assessment type": "session type",
    "class type": "session type",
    "cohort": "student groups",
    "cohort group": "student groups",
    "cohort groups": "student groups",
    "group": "student groups",
    "groups": "student groups",
    "section": "student groups",
    "student group": "student groups",
    "student groups": "student groups",
    "tutorial group": "student groups",
    "tutorial groups": "student groups",
    "type": "session type",
    "weekday": "day",
}


def _json(data: dict[str, Any]) -> str:
    """Return deterministic, Unicode-safe JSON for the model."""
    return json.dumps(data, ensure_ascii=False, indent=2, default=str)


def _error(code: str, message: str, **details: Any) -> str:
    payload: dict[str, Any] = {
        "status": "error",
        "error": {
            "code": code,
            "message": message,
        },
    }
    if details:
        payload["error"]["details"] = details
    return _json(payload)


def _is_inside(child: Path, parent: Path) -> bool:
    try:
        child.relative_to(parent)
        return True
    except ValueError:
        return False


def _available_files() -> list[dict[str, Any]]:
    if not FAKE_DATA_DIR.is_dir():
        return []

    files: list[dict[str, Any]] = []
    for path in sorted(FAKE_DATA_DIR.iterdir(), key=lambda item: item.name.casefold()):
        if (
            path.is_file()
            and not path.name.startswith("~$")
            and path.suffix.casefold() in SUPPORTED_EXTENSIONS
        ):
            files.append(
                {
                    "name": path.name,
                    "stem": path.stem,
                    "format": path.suffix.casefold().lstrip("."),
                    "size_bytes": path.stat().st_size,
                }
            )
    return files


def _resolve_requested_file(requested_value: str) -> tuple[Path | None, str | None]:
    """Resolve a user-supplied name while preventing access outside fake data."""
    requested_value = requested_value.strip().strip("\"'")
    if not requested_value:
        return None, _error(
            "missing_file_name",
            "A file name is required.",
            available_files=_available_files(),
        )

    requested = Path(requested_value)
    if requested.is_absolute():
        resolved = requested.resolve()
        if not _is_inside(resolved, FAKE_DATA_DIR):
            return None, _error(
                "path_outside_fake_data",
                "The tool can only read files inside the fake data folder.",
            )
        if resolved.is_file():
            return resolved, None
    else:
        parts = list(requested.parts)
        if parts and parts[0].casefold().replace("_", " ") == "fake data":
            parts = parts[1:]
        relative = Path(*parts) if parts else Path()
        candidate = (FAKE_DATA_DIR / relative).resolve()
        if not _is_inside(candidate, FAKE_DATA_DIR):
            return None, _error(
                "path_outside_fake_data",
                "The requested relative path leaves the fake data folder.",
            )
        if candidate.is_file():
            return candidate, None

    available_paths = [
        path
        for path in FAKE_DATA_DIR.iterdir()
        if path.is_file()
        and not path.name.startswith("~$")
        and path.suffix.casefold() in SUPPORTED_EXTENSIONS
    ]
    requested_name = requested.name.casefold()
    requested_stem = requested.stem.casefold()

    exact_matches = [
        path
        for path in available_paths
        if path.name.casefold() == requested_name
        or (not requested.suffix and path.stem.casefold() == requested_name)
    ]
    if len(exact_matches) == 1:
        return exact_matches[0].resolve(), None

    partial_matches = [
        path
        for path in available_paths
        if requested_stem and requested_stem in path.stem.casefold()
    ]
    if len(partial_matches) == 1:
        return partial_matches[0].resolve(), None
    if len(partial_matches) > 1:
        return None, _error(
            "ambiguous_file_name",
            "The requested name matches multiple files. Use an exact file name.",
            matches=[path.name for path in sorted(partial_matches)],
        )

    return None, _error(
        "file_not_found",
        "No matching Excel or PDF file was found in the fake data folder.",
        requested=requested_value,
        available_files=_available_files(),
    )


def _safe_value(value: Any) -> Any:
    """Convert spreadsheet values to compact JSON-safe values."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, (date, time)):
        return value.isoformat()
    if isinstance(value, (int, float, bool)):
        return value

    text = str(value).replace("\x00", "").strip()
    if len(text) > MAX_CELL_CHARS:
        return text[: MAX_CELL_CHARS - 1] + "…"
    return text


def _normalize_label(value: Any) -> str:
    text = str(value or "").strip().casefold()
    text = re.sub(r"[()]", "", text)
    text = re.sub(r"[/_\-]+", " ", text)
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def _query_terms(query: str | None) -> list[list[str]]:
    """Turn a natural query into required terms with small schedule synonyms."""
    if not query:
        return []

    raw_terms = [
        quoted or plain
        for quoted, plain in re.findall(r'"([^"]+)"|(\S+)', query)
    ]
    terms: list[list[str]] = []
    for raw_term in raw_terms:
        normalized = _normalize_label(raw_term)
        if not normalized or normalized in QUERY_STOP_WORDS:
            continue
        canonical = QUERY_SYNONYMS.get(normalized, normalized)
        alternatives = {canonical}
        if canonical.endswith("s") and len(canonical) > 3:
            alternatives.add(canonical[:-1])
        terms.append(sorted(alternatives))
    return terms


def _text_matches_query(text: str, query: str | None) -> bool:
    terms = _query_terms(query)
    if not terms:
        return True
    haystack = text.casefold()
    normalized_haystack = _normalize_label(text)
    return all(
        any(
            alternative in haystack or alternative in normalized_haystack
            for alternative in term
        )
        for term in terms
    )


def _row_matches(values: list[Any], query: str | None) -> bool:
    haystack = " | ".join(
        str(value) for value in values if value not in (None, "")
    )
    return _text_matches_query(haystack, query)


def _resolve_table_filters(
    headers: list[str],
    filters: dict[str, str | list[str]] | None,
    sheet_name: str,
) -> tuple[dict[str, str], list[str]]:
    if not filters:
        return {}, []

    normalized_headers = {
        _normalize_label(header): header for header in headers
    }
    resolved: dict[str, str] = {}
    unresolved: list[str] = []

    for requested_key in filters:
        normalized_key = _normalize_label(requested_key)
        direct = normalized_headers.get(normalized_key)
        if direct:
            resolved[requested_key] = direct
            continue

        canonical_key = FILTER_KEY_ALIASES.get(normalized_key, normalized_key)
        preferences = FILTER_HEADER_PREFERENCES.get(canonical_key, [])
        match = next(
            (
                normalized_headers[_normalize_label(candidate)]
                for candidate in preferences
                if _normalize_label(candidate) in normalized_headers
            ),
            None,
        )
        if match:
            resolved[requested_key] = match
        elif canonical_key == "day" and _normalize_label(sheet_name) in {
            "monday",
            "tuesday",
            "wednesday",
            "thursday",
            "friday",
            "saturday",
            "sunday",
        }:
            resolved[requested_key] = "__sheet_name__"
        else:
            unresolved.append(requested_key)
    return resolved, unresolved


def _filter_value_matches(
    actual: Any,
    expected: str | list[str],
) -> bool:
    expected_values = expected if isinstance(expected, list) else [expected]
    actual_text = str(actual or "").casefold()
    return any(
        str(expected_value).strip().casefold() in actual_text
        for expected_value in expected_values
    )


def _row_matches_filters(
    row_by_header: dict[str, Any],
    filters: dict[str, str | list[str]] | None,
    resolved_filters: dict[str, str],
    sheet_name: str,
) -> bool:
    if not filters:
        return True
    for requested_key, expected in filters.items():
        header = resolved_filters[requested_key]
        actual = sheet_name if header == "__sheet_name__" else row_by_header.get(header)
        if not _filter_value_matches(actual, expected):
            return False
    return True


def _unique_headers(raw_headers: list[Any]) -> list[str]:
    headers: list[str] = []
    counts: dict[str, int] = {}
    for index, value in enumerate(raw_headers, start=1):
        base = str(value).strip() if value not in (None, "") else f"column_{index}"
        counts[base] = counts.get(base, 0) + 1
        headers.append(base if counts[base] == 1 else f"{base}_{counts[base]}")
    return headers


class _ExtractionBudget:
    """Shared row/character budget across all returned tables or pages."""

    def __init__(self, max_rows: int, max_chars: int) -> None:
        self.rows_remaining = max_rows
        # Leave room for workbook metadata and JSON structure.
        self.chars_remaining = max(2_000, max_chars - 12_000)
        self.truncated = False

    def accept(self, item: Any) -> bool:
        estimated_chars = len(
            json.dumps(item, ensure_ascii=False, indent=2, default=str)
        )
        if self.rows_remaining <= 0 or estimated_chars > self.chars_remaining:
            self.truncated = True
            return False
        self.rows_remaining -= 1
        self.chars_remaining -= estimated_chars
        return True


def _resolve_sheet_names(
    workbook: Any,
    sheet_name: str | None,
    query: str | None,
    filters: dict[str, str | list[str]] | None,
) -> tuple[list[str], str | None]:
    names = list(workbook.sheetnames)
    if not sheet_name:
        if query or filters:
            return names, None
        preferred = next(
            (
                name
                for name in ("Overview", "Doctor Directory", "Semester Timetable", "Room Inventory")
                if name in names
            ),
            names[0] if names else None,
        )
        return ([preferred] if preferred else []), None

    matches = [name for name in names if name.casefold() == sheet_name.strip().casefold()]
    if len(matches) == 1:
        return matches, None
    return [], _error(
        "sheet_not_found",
        "The requested worksheet was not found.",
        requested_sheet=sheet_name,
        available_sheets=names,
    )


def _worksheet_index(workbook: Any) -> list[dict[str, Any]]:
    return [
        {
            "name": sheet.title,
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "table_names": list(sheet.tables.keys()),
        }
        for sheet in workbook.worksheets
    ]


def _extract_excel(
    path: Path,
    sheet_name: str | None,
    query: str | None,
    filters: dict[str, str | list[str]] | None,
    row_offset: int,
    max_rows: int,
    max_chars: int,
) -> str:
    workbook = load_workbook(path, read_only=False, data_only=True)
    selected_names, sheet_error = _resolve_sheet_names(
        workbook, sheet_name, query, filters
    )
    if sheet_error:
        workbook.close()
        return sheet_error

    budget = _ExtractionBudget(max_rows=max_rows, max_chars=max_chars)
    extracted_sheets: list[dict[str, Any]] = []
    total_matches_found = 0
    total_matches_returned = 0
    filter_tables_supported = 0
    filter_resolution_examples: list[dict[str, Any]] = []
    available_table_columns: list[dict[str, Any]] = []

    for name in selected_names:
        sheet = workbook[name]
        sheet_payload: dict[str, Any] = {
            "name": name,
            "dimensions": {
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
            },
            "tables": [],
        }

        # Capture the common title/subtitle convention without assuming it exists.
        context_rows: list[dict[str, Any]] = []
        for row_number in range(1, min(sheet.max_row, 3) + 1):
            values = [
                _safe_value(sheet.cell(row=row_number, column=column).value)
                for column in range(1, min(sheet.max_column, 20) + 1)
            ]
            nonempty = [value for value in values if value not in (None, "")]
            if nonempty:
                context_rows.append({"excel_row": row_number, "values": nonempty})
        if context_rows:
            sheet_payload["context"] = context_rows

        for table in sheet.tables.values():
            # openpyxl's type hints allow ``None`` for unbounded ranges, while an
            # Excel table always has a concrete rectangular reference.
            boundaries = range_boundaries(table.ref)
            if not all(isinstance(boundary, int) for boundary in boundaries):
                continue
            min_col, min_row, max_col, max_row = cast(
                tuple[int, int, int, int], boundaries
            )
            raw_headers = [
                sheet.cell(row=min_row, column=column).value
                for column in range(min_col, max_col + 1)
            ]
            headers = _unique_headers(raw_headers)
            returned_rows: list[dict[str, Any]] = []
            resolved_filters, unresolved_filters = _resolve_table_filters(
                headers, filters, name
            )
            available_table_columns.append(
                {
                    "sheet": name,
                    "table": table.name,
                    "columns": headers,
                }
            )
            if filters and unresolved_filters:
                continue
            if filters:
                filter_tables_supported += 1
                filter_resolution_examples.append(
                    {
                        "sheet": name,
                        "table": table.name,
                        "resolved_columns": resolved_filters,
                    }
                )

            for row_number in range(min_row + 1, max_row + 1):
                raw_values = [
                    sheet.cell(row=row_number, column=column).value
                    for column in range(min_col, max_col + 1)
                ]
                row_by_header = dict(zip(headers, raw_values))
                if not _row_matches(raw_values, query):
                    continue
                if not _row_matches_filters(
                    row_by_header,
                    filters,
                    resolved_filters,
                    name,
                ):
                    continue
                total_matches_found += 1
                if total_matches_found <= row_offset:
                    continue
                record = {
                    "excel_row": row_number,
                    "values": {
                        header: _safe_value(value)
                        for header, value in zip(headers, raw_values)
                    },
                }
                if not budget.accept(record):
                    continue
                returned_rows.append(record)
                total_matches_returned += 1

            if returned_rows or (not query and not filters) or (filters and not unresolved_filters):
                sheet_payload["tables"].append(
                    {
                        "name": table.name,
                        "range": table.ref,
                        "columns": headers,
                        "filters_resolved_to": resolved_filters or None,
                        "returned_row_count": len(returned_rows),
                        "rows": returned_rows,
                    }
                )

        # Raw-row fallback is only safe when structured column filters are absent.
        if not sheet.tables and not filters:
            raw_matches: list[dict[str, Any]] = []
            for row_number in range(1, sheet.max_row + 1):
                raw_values = [
                    sheet.cell(row=row_number, column=column).value
                    for column in range(1, sheet.max_column + 1)
                ]
                if not _row_matches(raw_values, query):
                    continue
                nonempty_cells = {
                    get_column_letter(column): _safe_value(value)
                    for column, value in enumerate(raw_values, start=1)
                    if value not in (None, "")
                }
                if not nonempty_cells:
                    continue
                total_matches_found += 1
                if total_matches_found <= row_offset:
                    continue
                record = {"excel_row": row_number, "cells": nonempty_cells}
                if not budget.accept(record):
                    continue
                raw_matches.append(record)
                total_matches_returned += 1
            if raw_matches:
                sheet_payload["raw_rows"] = raw_matches

        if (
            sheet_payload["tables"]
            or sheet_payload.get("raw_rows")
            or (not query and not filters)
        ):
            extracted_sheets.append(sheet_payload)

    if filters and filter_tables_supported == 0:
        workbook.close()
        return _error(
            "filters_not_resolved",
            "No worksheet table contains columns matching all requested filters.",
            requested_filters=filters,
            available_tables=available_table_columns[:50],
            hint=(
                "Use exact column headers from available_tables or inspect a "
                "specific worksheet first."
            ),
        )

    has_more = total_matches_found > row_offset + total_matches_returned
    truncated = budget.truncated or has_more

    payload = {
        "status": "ok",
        "file": {
            "name": path.name,
            "format": path.suffix.casefold().lstrip("."),
            "size_bytes": path.stat().st_size,
        },
        "workbook": {
            "sheet_count": len(workbook.sheetnames),
            "sheets": _worksheet_index(workbook),
        },
        "request": {
            "sheet_name": sheet_name,
            "query": query,
            "query_terms": _query_terms(query),
            "filters": filters,
            "row_offset": row_offset,
            "query_scope_warning": (
                (
                    "Query terms may match any column. Use structured filters "
                    "for precise or exhaustive multi-condition requests."
                )
                if query and not filters
                else None
            ),
        },
        "extraction": {
            "selected_sheets": selected_names,
            "matching_rows_found": total_matches_found,
            "matching_rows_returned": total_matches_returned,
            "matching_rows_skipped_by_offset": min(
                row_offset, total_matches_found
            ),
            "has_more": has_more,
            "next_row_offset": (
                row_offset + total_matches_returned
                if has_more and total_matches_returned
                else None
            ),
            "sheets": extracted_sheets,
        },
        "filter_diagnostics": {
            "tables_supporting_all_filters": filter_tables_supported,
            "resolution_examples": filter_resolution_examples[:20],
        },
        "limits": {
            "max_rows": max_rows,
            "max_chars": max_chars,
            "truncated": truncated,
            "note": (
                (
                    "Call again with next_row_offset to retrieve remaining "
                    "matches, or narrow the filters."
                )
                if has_more
                else (
                    "Increase max_chars or narrow the filters."
                    if budget.truncated
                    else None
                )
            ),
            "completeness": (
                "complete"
                if not truncated
                else "partial"
            ),
        },
    }
    workbook.close()
    return _json(payload)


def _clean_pdf_text(text: str) -> str:
    text = text.replace("\x00", "")
    text = "\n".join(line.rstrip() for line in text.splitlines())
    return re.sub(r"\n{3,}", "\n\n", text).strip()


def _query_snippet(text: str, query: str, radius: int = 800) -> str:
    position = text.casefold().find(query.casefold())
    if position < 0:
        return text
    start = max(0, position - radius)
    end = min(len(text), position + len(query) + radius)
    prefix = "…" if start else ""
    suffix = "…" if end < len(text) else ""
    return prefix + text[start:end].strip() + suffix


def _extract_pdf(
    path: Path,
    page_number: int | None,
    query: str | None,
    max_pages: int,
    max_chars: int,
) -> str:
    reader = PdfReader(str(path))
    if reader.is_encrypted and reader.decrypt("") == 0:
        return _error(
            "encrypted_pdf",
            "The PDF is encrypted and cannot be read without a password.",
            file=path.name,
        )

    page_count = len(reader.pages)
    if page_number is not None:
        if page_number < 1 or page_number > page_count:
            return _error(
                "page_out_of_range",
                "The requested PDF page is outside the document.",
                requested_page=page_number,
                page_count=page_count,
            )
        page_indexes = [page_number - 1]
    else:
        page_indexes = list(range(page_count))

    pages: list[dict[str, Any]] = []
    chars_remaining = max(1_000, max_chars - 4_000)
    truncated = False
    matched_pages = 0

    for page_index in page_indexes:
        text = _clean_pdf_text(reader.pages[page_index].extract_text() or "")
        if query and not _text_matches_query(text, query):
            continue
        if query:
            terms = _query_terms(query)
            snippet_term = terms[0][0] if terms else query
            text = _query_snippet(text, snippet_term)
        matched_pages += 1

        if len(pages) >= max_pages:
            truncated = True
            break
        if len(text) > chars_remaining:
            text = text[: max(0, chars_remaining - 1)] + "…"
            truncated = True
        pages.append(
            {
                "page_number": page_index + 1,
                "text": text,
            }
        )
        chars_remaining -= len(text)
        if chars_remaining <= 0:
            truncated = True
            break

    metadata = reader.metadata or {}
    payload = {
        "status": "ok",
        "file": {
            "name": path.name,
            "format": "pdf",
            "size_bytes": path.stat().st_size,
        },
        "pdf": {
            "page_count": page_count,
            "metadata": {
                "title": _safe_value(metadata.get("/Title")),
                "author": _safe_value(metadata.get("/Author")),
                "subject": _safe_value(metadata.get("/Subject")),
            },
        },
        "request": {
            "page_number": page_number,
            "query": query,
        },
        "extraction": {
            "matched_page_count": matched_pages,
            "pages": pages,
        },
        "limits": {
            "max_pages": max_pages,
            "max_chars": max_chars,
            "truncated": truncated,
            "note": (
                "Request a specific page or add a query to narrow the result."
                if truncated
                else None
            ),
        },
    }
    return _json(payload)


@tool
def get_schedule(
    uploaded_file_path: str,
    sheet_name: str | None = None,
    query: str | None = None,
    filters: dict[str, str | list[str]] | None = None,
    row_offset: int = 0,
    page_number: int | None = None,
    max_rows: int = DEFAULT_MAX_ROWS,
    max_pages: int = DEFAULT_MAX_PAGES,
    max_chars: int = DEFAULT_MAX_CHARS,
) -> str:
    """Extract AI-ready data from an Excel or PDF file in the fake data folder.

    Set ``uploaded_file_path`` to the exact schedule file name when possible.
    Never infer a file's contents from its name. If the correct Excel worksheet
    is unknown, call with the file name first and use the returned workbook
    discovery/index before selecting ``sheet_name``.

    Use ``filters`` for precise or exhaustive Excel retrieval. Filters use AND
    logic and may reference exact uploaded headers or supported canonical fields:
    ``student_groups``, ``session_type``, ``day``, ``major``, ``course``,
    ``room``, ``instructor``, ``period``, and ``week``. ``query`` is a
    natural-text row search and should not replace column filters when several
    exact conditions must all match. Example:
    ``filters={"student_groups": "ARC", "session_type": "Tutorial",
    "day": "Sunday"}``.

    For an exhaustive request, verify
    ``extraction.matching_rows_found == extraction.matching_rows_returned`` and
    ``extraction.has_more == false``. When ``extraction.has_more`` is true, call
    again with ``row_offset`` set to ``extraction.next_row_offset`` until all
    pages are retrieved. Never describe a partial page as the complete result.

    For PDFs, ``page_number`` is one-based. Use a page number for a known page or
    ``query`` to locate relevant text. Respect row, page, and character limits.
    A response marked truncated, partial, or error is incomplete data: narrow the
    request or paginate, and never invent the missing content.
    """
    if not FAKE_DATA_DIR.is_dir():
        return _error(
            "fake_data_folder_missing",
            "The fake data folder does not exist.",
            expected_path=str(FAKE_DATA_DIR),
        )
    print("getScheduleUSED")
    path, resolution_error = _resolve_requested_file(uploaded_file_path)
    if resolution_error:
        return resolution_error
    assert path is not None

    if path.name.startswith("~$"):
        return _error(
            "temporary_office_file",
            "Temporary Office lock files cannot be read. Request the real workbook instead.",
            file=path.name,
        )

    extension = path.suffix.casefold()
    if extension not in SUPPORTED_EXTENSIONS:
        return _error(
            "unsupported_format",
            "Only .xlsx, .xlsm, and .pdf files are supported.",
            file=path.name,
            extension=extension,
        )

    max_rows = max(1, min(int(max_rows), MAX_ROW_LIMIT))
    max_pages = max(1, min(int(max_pages), MAX_PAGE_LIMIT))
    max_chars = max(5_000, min(int(max_chars), MAX_CHAR_LIMIT))
    row_offset = max(0, int(row_offset))
    sheet_name = sheet_name.strip() if sheet_name and sheet_name.strip() else None
    query = query.strip() if query and query.strip() else None
    filters = filters or None
    if filters:
        cleaned_filters: dict[str, str | list[str]] = {}
        for key, value in filters.items():
            cleaned_key = str(key).strip()
            if not cleaned_key:
                return _error(
                    "invalid_filter",
                    "Filter column names cannot be empty.",
                )
            if isinstance(value, list):
                cleaned_values = [
                    str(item).strip()
                    for item in value
                    if str(item).strip()
                ]
                if not cleaned_values:
                    return _error(
                        "invalid_filter",
                        "Filter value lists cannot be empty.",
                        filter=cleaned_key,
                    )
                cleaned_filters[cleaned_key] = cleaned_values
            else:
                cleaned_value = str(value).strip()
                if not cleaned_value:
                    return _error(
                        "invalid_filter",
                        "Filter values cannot be empty.",
                        filter=cleaned_key,
                    )
                cleaned_filters[cleaned_key] = cleaned_value
        filters = cleaned_filters

    try:
        if extension in {".xlsx", ".xlsm"}:
            if page_number is not None:
                return _error(
                    "page_not_applicable",
                    "page_number is only valid for PDF files.",
                    file=path.name,
                )
            return _extract_excel(
                path=path,
                sheet_name=sheet_name,
                query=query,
                filters=filters,
                row_offset=row_offset,
                max_rows=max_rows,
                max_chars=max_chars,
            )

        if filters is not None:
            return _error(
                "filters_not_applicable",
                "Structured column filters are only valid for Excel files.",
                file=path.name,
            )
        if row_offset:
            return _error(
                "row_offset_not_applicable",
                "row_offset is only valid for Excel files.",
                file=path.name,
            )
        if sheet_name is not None:
            return _error(
                "sheet_not_applicable",
                "sheet_name is only valid for Excel files.",
                file=path.name,
            )
        return _extract_pdf(
            path=path,
            page_number=page_number,
            query=query,
            max_pages=max_pages,
            max_chars=max_chars,
        )
    except PermissionError:
        return _error(
            "file_locked",
            "The file is open or locked by another process. Close it and try again.",
            file=path.name,
        )
    except Exception as exc:  # Tool responses must remain structured and UI-safe.
        return _error(
            "extraction_failed",
            "The schedule file could not be extracted.",
            file=path.name,
            exception_type=type(exc).__name__,
            reason=str(exc),
        )
