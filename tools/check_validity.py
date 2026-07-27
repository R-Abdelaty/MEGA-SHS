"""Hybrid schedule normalization and deterministic validity checks.

The AI is responsible for confirming unfamiliar input layouts. This module is
responsible for exhaustively applying the confirmed mappings to every row and
running repeatable conflict checks.
"""

from __future__ import annotations

import json
import re
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Iterable

from langchain.tools import tool
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from .get_schedule import FAKE_DATA_DIR


SUPPORTED_EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}
DAY_NAMES = {
    "monday",
    "tuesday",
    "wednesday",
    "thursday",
    "friday",
    "saturday",
    "sunday",
}
DEFAULT_RULES: dict[str, Any] = {
    "teaching_weeks": list(range(1, 13)),
    "final_exam_start_times": ["09:00", "14:00"],
    "preferred_quiz_periods": ["1", "5"],
    "quiz_period_rule_is_hard": False,
}
ALLOWED_ROLES = {
    "sessions",
    "doctor_sessions",
    "rooms",
    "room_availability",
    "periods",
    "staff_directory",
    "ignore",
}

# Canonical fields are deliberately independent from the fake-data headers.
# Exact normalized aliases are safe to auto-map. Anything else is returned to
# the agent as a mapping request instead of being guessed.
FIELD_ALIASES: dict[str, set[str]] = {
    "session_id": {
        "session id",
        "booking id",
        "assessment id",
        "exam id",
        "event id",
        "class id",
        "reservation id",
    },
    "session_type": {
        "session type",
        "activity type",
        "assessment type",
        "exam type",
        "class type",
        "type of session",
    },
    "course_id": {
        "course id",
        "course code",
        "subject code",
        "module code",
        "module id",
    },
    "course_name": {
        "course name",
        "subject",
        "subject name",
        "module",
        "module name",
        "course event",
    },
    "majors": {
        "major",
        "majors",
        "major program",
        "majors covered",
        "program",
        "programs",
    },
    "major_codes": {
        "major code",
        "major codes",
        "program code",
        "program codes",
    },
    "student_groups": {
        "cohort groups",
        "cohort group",
        "student groups",
        "student group",
        "tutorial group",
        "tutorial groups",
        "t group",
        "t groups",
        "group id",
        "section",
    },
    "instructor": {
        "instructor",
        "lecturer",
        "doctor",
        "doctor name",
        "staff",
        "staff name",
        "teacher",
        "ta",
    },
    "staff_id": {
        "staff id",
        "doctor id",
        "lecturer id",
        "instructor id",
        "employee id",
    },
    "room_id": {
        "room",
        "room id",
        "room code",
        "venue",
        "location",
        "hall",
        "lab",
        "room assignments students",
    },
    "room_type": {
        "room type",
        "venue type",
        "location type",
        "lab profile",
        "profile",
        "type",
    },
    "room_capacity": {
        "room capacity",
        "capacity",
        "maximum capacity",
        "max capacity",
        "seats",
    },
    "expected_students": {
        "expected students",
        "student count",
        "students",
        "enrollment",
        "enrolment",
        "class size",
        "attendees",
    },
    "day": {
        "day",
        "weekday",
        "day of week",
    },
    "date": {
        "date",
        "session date",
        "exam date",
        "assessment date",
    },
    "week": {
        "week",
        "weeks",
        "teaching week",
        "semester week",
    },
    "period": {
        "period",
        "period id",
        "slot",
        "time slot",
        "timeslot",
    },
    "start": {
        "start",
        "start time",
        "from",
        "begins",
    },
    "end": {
        "end",
        "end time",
        "to",
        "finishes",
    },
    "status": {
        "status",
        "availability",
        "booking status",
    },
}

REQUIRED_SESSION_FIELDS = {
    "course": ("course_id", "course_name"),
    "day_or_date": ("day", "date"),
    "time": ("start", "period"),
}


def _json(data: dict[str, Any]) -> str:
    return json.dumps(data, ensure_ascii=False, indent=2, default=str)


def _error(code: str, message: str, **details: Any) -> str:
    payload: dict[str, Any] = {
        "status": "error",
        "error": {"code": code, "message": message},
    }
    if details:
        payload["error"]["details"] = details
    return _json(payload)


def _normalize_label(value: Any) -> str:
    text = str(value or "").strip().casefold()
    text = re.sub(r"[()]", "", text)
    text = re.sub(r"[/_\-]+", " ", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


NORMALIZED_ALIASES = {
    field_name: {_normalize_label(alias) for alias in aliases}
    for field_name, aliases in FIELD_ALIASES.items()
}
HEADER_PREFERENCES: dict[str, list[str]] = {
    "student_groups": [
        "cohort groups",
        "cohort group",
        "student groups",
        "student group",
        "tutorial groups",
        "tutorial group",
        "t groups",
        "t group",
        "group id",
        "section",
    ],
    "period": ["period id", "period", "time slot", "timeslot", "slot"],
    "course_id": [
        "course id",
        "course code",
        "module id",
        "module code",
        "subject code",
    ],
}


def _safe_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, (date, time)):
        return value.isoformat()
    if isinstance(value, (int, float, bool)):
        return value
    return str(value).replace("\x00", "").strip()


def _is_inside(child: Path, parent: Path) -> bool:
    try:
        child.relative_to(parent)
        return True
    except ValueError:
        return False


def _resolve_file(file_name: str) -> tuple[Path | None, dict[str, Any] | None]:
    requested = Path(str(file_name).strip().strip("\"'"))
    if not requested.name:
        return None, {
            "file": file_name,
            "code": "missing_file_name",
            "message": "A schedule file name is required.",
        }

    candidate = requested if requested.is_absolute() else FAKE_DATA_DIR / requested
    candidate = candidate.resolve()
    if not _is_inside(candidate, FAKE_DATA_DIR):
        return None, {
            "file": file_name,
            "code": "path_outside_fake_data",
            "message": "Validation can only read files inside the fake data folder.",
        }
    if candidate.name.startswith("~$"):
        return None, {
            "file": file_name,
            "code": "temporary_office_file",
            "message": "An Office lock file is not a schedule source.",
        }
    if not candidate.is_file():
        matches = [
            path
            for path in FAKE_DATA_DIR.iterdir()
            if path.is_file()
            and not path.name.startswith("~$")
            and (
                path.name.casefold() == requested.name.casefold()
                or (
                    not requested.suffix
                    and path.stem.casefold() == requested.name.casefold()
                )
            )
        ]
        if len(matches) == 1:
            candidate = matches[0].resolve()
        else:
            return None, {
                "file": file_name,
                "code": "file_not_found",
                "message": "The requested schedule file was not found.",
            }
    return candidate, None


def _source_key(file_name: str, sheet_name: str, region_name: str) -> str:
    return f"{file_name}::{sheet_name}::{region_name}"


def _select_override(
    overrides: dict[str, dict[str, str]],
    file_name: str,
    sheet_name: str,
    region_name: str,
) -> dict[str, str]:
    keys = (
        _source_key(file_name, sheet_name, region_name),
        f"{file_name}::{sheet_name}",
        file_name,
    )
    for key in keys:
        value = overrides.get(key)
        if isinstance(value, dict):
            return {str(k): str(v) for k, v in value.items()}
    return {}


def _auto_mapping(headers: list[str]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    normalized_headers = [(_normalize_label(header), header) for header in headers]
    for field_name, aliases in NORMALIZED_ALIASES.items():
        matches = [
            (normalized, header)
            for normalized, header in normalized_headers
            if normalized in aliases
        ]
        if not matches:
            continue
        preferences = [
            _normalize_label(value)
            for value in HEADER_PREFERENCES.get(field_name, [])
        ]
        matches.sort(
            key=lambda item: (
                preferences.index(item[0]) if item[0] in preferences else len(preferences),
                headers.index(item[1]),
            )
        )
        mapping[field_name] = matches[0][1]
    return mapping


def _merge_mapping(
    headers: list[str], override: dict[str, str]
) -> tuple[dict[str, str], list[dict[str, str]]]:
    mapping = _auto_mapping(headers)
    problems: list[dict[str, str]] = []
    header_lookup = {_normalize_label(header): header for header in headers}

    for canonical, uploaded_header in override.items():
        if canonical.startswith("_"):
            continue
        if canonical not in FIELD_ALIASES:
            problems.append(
                {
                    "field": canonical,
                    "reason": "Unknown canonical field.",
                }
            )
            continue
        actual_header = header_lookup.get(_normalize_label(uploaded_header))
        if actual_header is None:
            problems.append(
                {
                    "field": canonical,
                    "reason": f"Header {uploaded_header!r} was not found.",
                }
            )
            continue
        mapping[canonical] = actual_header
    return mapping, problems


def _sheet_default_day(sheet_name: str) -> str | None:
    normalized = _normalize_label(sheet_name)
    return normalized.title() if normalized in DAY_NAMES else None


def _infer_role(
    file_name: str,
    sheet_name: str,
    mapping: dict[str, str],
    override_role: str | None,
) -> tuple[str, str]:
    if override_role:
        normalized_role = override_role.strip().casefold()
        if normalized_role in ALLOWED_ROLES:
            return normalized_role, "confirmed_by_mapping"
        return "ignore", "invalid_override_role"

    context = _normalize_label(f"{file_name} {sheet_name}")
    sheet_context = _normalize_label(sheet_name)
    fields = set(mapping)
    has_time = bool({"start", "period"} & fields)
    has_course = bool({"course_id", "course_name"} & fields)
    has_day = bool({"day", "date"} & fields) or _sheet_default_day(sheet_name) is not None

    if "period" in context and {"start", "end", "period"} <= fields:
        return "periods", "recognized_period_table"
    if (
        ("room inventory" in context or "room directory" in context or "lab directory" in context)
        and "room_id" in fields
        and "room_capacity" in fields
    ):
        return "rooms", "recognized_room_inventory"
    if "doctor directory" in context:
        return "staff_directory", "recognized_doctor_directory"
    if "room schedule" in context and "exam exceptions" in context:
        return "room_availability", "recognized_room_exam_reservations"
    if "room schedule" in context and {"room_id", "status"} <= fields and has_day and has_time:
        return "room_availability", "recognized_room_availability"
    if "doctor schedule" in context and has_course and has_day and has_time:
        return "doctor_sessions", "recognized_doctor_schedule"
    if any(
        phrase in sheet_context
        for phrase in (
            "assumptions",
            "overview",
            "validation",
            "coverage",
            "blackout",
            "time rules",
            "maintenance",
        )
    ):
        return "ignore", "recognized_metadata_or_policy_table"
    if has_course and has_day and has_time:
        return "sessions", "recognized_session_fields"
    if has_course and (has_day or has_time):
        return "sessions", "possible_session_table_requires_mapping"
    if (
        any(word in context for word in ("schedule", "timetable", "calendar"))
        and fields
    ):
        return "sessions", "possible_schedule_layout_requires_mapping"
    if "room" in context and "room_id" in fields:
        return "rooms", "possible_room_inventory_requires_mapping"
    return "ignore", "not_a_recognized_validation_table"


@dataclass
class Region:
    file_name: str
    sheet_name: str
    name: str
    headers: list[str]
    rows: list[tuple[int, dict[str, Any]]]
    mapping: dict[str, str]
    role: str
    role_reason: str
    mapping_problems: list[dict[str, str]] = field(default_factory=list)

    @property
    def key(self) -> str:
        return _source_key(self.file_name, self.sheet_name, self.name)


def _table_rows(sheet: Any, table: Any) -> tuple[list[str], list[tuple[int, dict[str, Any]]]]:
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    raw_headers = [
        sheet.cell(row=min_row, column=column).value
        for column in range(min_col, max_col + 1)
    ]
    headers = [
        str(value).strip() if value not in (None, "") else f"Column {index}"
        for index, value in enumerate(raw_headers, start=1)
    ]
    rows: list[tuple[int, dict[str, Any]]] = []
    for row_number in range(min_row + 1, max_row + 1):
        values = [
            sheet.cell(row=row_number, column=column).value
            for column in range(min_col, max_col + 1)
        ]
        if not any(value not in (None, "") for value in values):
            continue
        rows.append((row_number, dict(zip(headers, values))))
    return headers, rows


def _detected_rows(
    sheet: Any,
) -> tuple[str, list[str], list[tuple[int, dict[str, Any]]]] | None:
    best_row = 0
    best_score = 0
    best_values: list[Any] = []
    for row_number in range(1, min(sheet.max_row, 30) + 1):
        values = [
            sheet.cell(row=row_number, column=column).value
            for column in range(1, sheet.max_column + 1)
        ]
        normalized = {_normalize_label(value) for value in values if value not in (None, "")}
        score = sum(
            1
            for aliases in NORMALIZED_ALIASES.values()
            if normalized.intersection(aliases)
        )
        if score > best_score:
            best_row = row_number
            best_score = score
            best_values = values
    if best_score < 3:
        return None

    last_column = max(
        index
        for index, value in enumerate(best_values, start=1)
        if value not in (None, "")
    )
    headers = [
        str(value).strip() if value not in (None, "") else f"Column {index}"
        for index, value in enumerate(best_values[:last_column], start=1)
    ]
    rows: list[tuple[int, dict[str, Any]]] = []
    for row_number in range(best_row + 1, sheet.max_row + 1):
        values = [
            sheet.cell(row=row_number, column=column).value
            for column in range(1, last_column + 1)
        ]
        if not any(value not in (None, "") for value in values):
            continue
        rows.append((row_number, dict(zip(headers, values))))
    return f"DetectedRange@{best_row}", headers, rows


def _discover_regions(
    path: Path,
    overrides: dict[str, dict[str, str]],
) -> list[Region]:
    workbook = load_workbook(path, read_only=False, data_only=True)
    regions: list[Region] = []
    try:
        for sheet in workbook.worksheets:
            discovered: list[tuple[str, list[str], list[tuple[int, dict[str, Any]]]]] = []
            if sheet.tables:
                for table in sheet.tables.values():
                    headers, rows = _table_rows(sheet, table)
                    discovered.append((table.name, headers, rows))
            else:
                detected = _detected_rows(sheet)
                if detected:
                    discovered.append(detected)

            for region_name, headers, rows in discovered:
                override = _select_override(
                    overrides, path.name, sheet.title, region_name
                )
                mapping, mapping_problems = _merge_mapping(headers, override)
                if (
                    override.get("_role")
                    and override["_role"].strip().casefold() not in ALLOWED_ROLES
                ):
                    mapping_problems.append(
                        {
                            "field": "_role",
                            "reason": (
                                f"Unknown role {override['_role']!r}. Expected one "
                                f"of {sorted(ALLOWED_ROLES)}."
                            ),
                        }
                    )
                role, role_reason = _infer_role(
                    path.name,
                    sheet.title,
                    mapping,
                    override.get("_role"),
                )
                regions.append(
                    Region(
                        file_name=path.name,
                        sheet_name=sheet.title,
                        name=region_name,
                        headers=headers,
                        rows=rows,
                        mapping=mapping,
                        role=role,
                        role_reason=role_reason,
                        mapping_problems=mapping_problems,
                    )
                )
    finally:
        workbook.close()
    return regions


def _mapped_value(row: dict[str, Any], mapping: dict[str, str], field_name: str) -> Any:
    header = mapping.get(field_name)
    return row.get(header) if header else None


def _as_text(value: Any) -> str | None:
    if value in (None, ""):
        return None
    return str(value).strip()


def _as_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(value)
    match = re.search(r"-?\d+", str(value))
    return int(match.group()) if match else None


def _time_minutes(value: Any) -> int | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.hour * 60 + value.minute
    if isinstance(value, time):
        return value.hour * 60 + value.minute
    if isinstance(value, (int, float)) and 0 <= float(value) < 1:
        return round(float(value) * 24 * 60)

    text = str(value).strip()
    for pattern in ("%H:%M:%S", "%H:%M", "%I:%M %p", "%I %p"):
        try:
            parsed = datetime.strptime(text, pattern)
            return parsed.hour * 60 + parsed.minute
        except ValueError:
            continue
    return None


def _format_minutes(value: int | None) -> str | None:
    if value is None:
        return None
    return f"{value // 60:02d}:{value % 60:02d}"


def _normalize_day(value: Any) -> str | None:
    if value in (None, ""):
        return None
    normalized = _normalize_label(value)
    for day_name in DAY_NAMES:
        if normalized == day_name or normalized.startswith(day_name[:3]):
            return day_name.title()
    return str(value).strip()


def _normalize_date(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, pattern).date().isoformat()
        except ValueError:
            continue
    return text


def _split_values(value: Any, split_commas: bool = True) -> list[str]:
    if value in (None, ""):
        return []
    pattern = r"[;|\n,]+" if split_commas else r"[;|\n]+"
    return [
        part.strip()
        for part in re.split(pattern, str(value))
        if part.strip()
    ]


def _room_values(value: Any) -> list[str]:
    rooms: list[str] = []
    for part in _split_values(value):
        cleaned = re.sub(r"\s*\([^)]*\)\s*$", "", part).strip()
        if cleaned:
            rooms.append(cleaned)
    return rooms


def _week_values(value: Any, default_weeks: list[int]) -> set[int]:
    if value in (None, ""):
        return set(default_weeks)
    if isinstance(value, (int, float)):
        return {int(value)}

    weeks: set[int] = set()
    for start_text, end_text in re.findall(
        r"(\d+)(?:\s*[-–]\s*(\d+))?", str(value)
    ):
        start = int(start_text)
        end = int(end_text) if end_text else start
        if start <= end and end - start <= 60:
            weeks.update(range(start, end + 1))
    return weeks or set(default_weeks)


def _session_type(row: dict[str, Any], mapping: dict[str, str]) -> str:
    value = _as_text(_mapped_value(row, mapping, "session_type"))
    return value.casefold() if value else "unspecified"


def _period_key(value: Any) -> str | None:
    if value in (None, ""):
        return None
    text = str(value).strip()
    match = re.search(r"\d+", text)
    return match.group() if match else text.casefold()


def _build_period_map(regions: list[Region]) -> dict[str, tuple[int, int]]:
    period_map: dict[str, tuple[int, int]] = {}
    for region in regions:
        if region.role != "periods":
            continue
        for _, row in region.rows:
            key = _period_key(_mapped_value(row, region.mapping, "period"))
            start = _time_minutes(_mapped_value(row, region.mapping, "start"))
            end = _time_minutes(_mapped_value(row, region.mapping, "end"))
            if key and start is not None and end is not None:
                period_map[key] = (start, end)
    return period_map


def _build_staff_directory(regions: list[Region]) -> dict[str, str]:
    staff: dict[str, str] = {}
    for region in regions:
        if region.role != "staff_directory":
            continue
        for _, row in region.rows:
            staff_id = _as_text(
                _mapped_value(row, region.mapping, "staff_id")
            )
            staff_name = _as_text(
                _mapped_value(row, region.mapping, "instructor")
            )
            if staff_id and staff_name:
                staff[staff_id.casefold()] = staff_name
    return staff


def _session_missing_mapping(region: Region) -> list[str]:
    missing: list[str] = []
    fields = set(region.mapping)
    if not fields.intersection(REQUIRED_SESSION_FIELDS["course"]):
        missing.append("course_id or course_name")
    if (
        not fields.intersection(REQUIRED_SESSION_FIELDS["day_or_date"])
        and _sheet_default_day(region.sheet_name) is None
    ):
        missing.append("day or date")
    if not fields.intersection(REQUIRED_SESSION_FIELDS["time"]):
        missing.append("start or period")
    return missing


def _room_missing_mapping(region: Region) -> list[str]:
    missing: list[str] = []
    if "room_id" not in region.mapping:
        missing.append("room_id")
    if "room_capacity" not in region.mapping:
        missing.append("room_capacity")
    return missing


def _availability_missing_mapping(region: Region) -> list[str]:
    missing: list[str] = []
    fields = set(region.mapping)
    if "session_id" not in fields:
        missing.append("session_id")
    if "room_id" not in fields:
        missing.append("room_id")
    if (
        not fields.intersection(REQUIRED_SESSION_FIELDS["day_or_date"])
        and _sheet_default_day(region.sheet_name) is None
    ):
        missing.append("day or date")
    if not fields.intersection(REQUIRED_SESSION_FIELDS["time"]):
        missing.append("start or period")
    return missing


def _staff_directory_missing_mapping(region: Region) -> list[str]:
    missing: list[str] = []
    if "staff_id" not in region.mapping:
        missing.append("staff_id")
    if "instructor" not in region.mapping:
        missing.append("instructor")
    return missing


@dataclass
class Session:
    source: dict[str, Any]
    session_id: str | None
    session_type: str
    course_id: str | None
    course_name: str | None
    majors: list[str]
    major_codes: list[str]
    student_groups: list[str]
    instructors: list[str]
    rooms: list[str]
    day: str | None
    date: str | None
    weeks: set[int]
    period: str | None
    start: int | None
    end: int | None
    expected_students: int | None
    inline_room_capacity: int | None

    def evidence(self) -> dict[str, Any]:
        return {
            **self.source,
            "session_id": self.session_id,
            "course_id": self.course_id,
            "course_name": self.course_name,
            "majors": self.majors,
            "major_codes": self.major_codes,
            "session_type": self.session_type,
            "day": self.day,
            "date": self.date,
            "weeks": sorted(self.weeks),
            "start": _format_minutes(self.start),
            "end": _format_minutes(self.end),
            "rooms": self.rooms,
            "instructors": self.instructors,
            "student_groups": self.student_groups,
        }


def _normalize_session(
    region: Region,
    row_number: int,
    row: dict[str, Any],
    period_map: dict[str, tuple[int, int]],
    staff_directory: dict[str, str],
    rules: dict[str, Any],
) -> Session:
    period = _period_key(_mapped_value(row, region.mapping, "period"))
    start = _time_minutes(_mapped_value(row, region.mapping, "start"))
    end = _time_minutes(_mapped_value(row, region.mapping, "end"))
    if period and period in period_map:
        start = start if start is not None else period_map[period][0]
        end = end if end is not None else period_map[period][1]

    instructor_value = _mapped_value(row, region.mapping, "instructor")
    if region.role == "doctor_sessions" and instructor_value in (None, ""):
        instructor_value = staff_directory.get(
            region.sheet_name.casefold(), region.sheet_name
        )

    day = _normalize_day(_mapped_value(row, region.mapping, "day"))
    day = day or _sheet_default_day(region.sheet_name)
    date_value = _normalize_date(_mapped_value(row, region.mapping, "date"))

    return Session(
        source={
            "file": region.file_name,
            "sheet": region.sheet_name,
            "region": region.name,
            "row": row_number,
        },
        session_id=_as_text(_mapped_value(row, region.mapping, "session_id")),
        session_type=_session_type(row, region.mapping),
        course_id=_as_text(_mapped_value(row, region.mapping, "course_id")),
        course_name=_as_text(_mapped_value(row, region.mapping, "course_name")),
        majors=_split_values(_mapped_value(row, region.mapping, "majors")),
        major_codes=_split_values(
            _mapped_value(row, region.mapping, "major_codes")
        ),
        student_groups=_split_values(
            _mapped_value(row, region.mapping, "student_groups")
        ),
        instructors=_split_values(instructor_value, split_commas=False),
        rooms=_room_values(_mapped_value(row, region.mapping, "room_id")),
        day=day,
        date=date_value,
        weeks=_week_values(
            _mapped_value(row, region.mapping, "week"),
            [int(value) for value in rules["teaching_weeks"]],
        ),
        period=period,
        start=start,
        end=end,
        expected_students=_as_int(
            _mapped_value(row, region.mapping, "expected_students")
        ),
        inline_room_capacity=_as_int(
            _mapped_value(row, region.mapping, "room_capacity")
        ),
    )


def _build_room_inventory(regions: list[Region]) -> dict[str, dict[str, Any]]:
    rooms: dict[str, dict[str, Any]] = {}
    for region in regions:
        if region.role != "rooms":
            continue
        for row_number, row in region.rows:
            room_id = _as_text(_mapped_value(row, region.mapping, "room_id"))
            if not room_id:
                continue
            rooms[room_id.casefold()] = {
                "room_id": room_id,
                "capacity": _as_int(
                    _mapped_value(row, region.mapping, "room_capacity")
                ),
                "room_type": _as_text(
                    _mapped_value(row, region.mapping, "room_type")
                ),
                "source": {
                    "file": region.file_name,
                    "sheet": region.sheet_name,
                    "region": region.name,
                    "row": row_number,
                },
            }
    return rooms


class IssueCollector:
    def __init__(self, max_issues: int) -> None:
        self.max_issues = max_issues
        self.items: list[dict[str, Any]] = []
        self.total = 0
        self.counts: dict[str, int] = defaultdict(int)
        self.severity_counts: dict[str, int] = defaultdict(int)

    def add(
        self,
        severity: str,
        code: str,
        message: str,
        evidence: list[dict[str, Any]],
        **details: Any,
    ) -> None:
        self.total += 1
        self.counts[code] += 1
        self.severity_counts[severity] += 1
        if len(self.items) < self.max_issues:
            issue: dict[str, Any] = {
                "severity": severity,
                "code": code,
                "message": message,
                "evidence": evidence,
            }
            if details:
                issue["details"] = details
            self.items.append(issue)


def _same_occurrence(left: Session, right: Session) -> bool:
    if left.date and right.date:
        return left.date == right.date
    if left.date or right.date:
        return False
    return bool(
        left.day
        and right.day
        and left.day.casefold() == right.day.casefold()
        and left.weeks.intersection(right.weeks)
    )


def _overlap(left: Session, right: Session) -> bool:
    return bool(
        left.start is not None
        and left.end is not None
        and right.start is not None
        and right.end is not None
        and left.start < right.end
        and right.start < left.end
    )


def _same_session(left: Session, right: Session) -> bool:
    if left.session_id and right.session_id:
        return left.session_id.casefold() == right.session_id.casefold()
    return False


def _check_grouped_overlaps(
    sessions: list[Session],
    attribute: str,
    code: str,
    label: str,
    collector: IssueCollector,
) -> tuple[int, int]:
    grouped: dict[str, list[Session]] = defaultdict(list)
    usable_sessions = 0
    for session in sessions:
        values = getattr(session, attribute)
        if values and session.start is not None and session.end is not None:
            usable_sessions += 1
        for value in values:
            grouped[value.casefold()].append(session)

    issue_count_before = collector.counts.get(code, 0)
    seen_pairs: set[tuple[str, int, str, int, str]] = set()
    for normalized_value, values in grouped.items():
        ordered = sorted(
            values,
            key=lambda item: (
                item.date or "",
                item.day or "",
                item.start if item.start is not None else -1,
            ),
        )
        for index, left in enumerate(ordered):
            for right in ordered[index + 1 :]:
                if left.start is None or left.end is None:
                    continue
                if right.start is None or right.end is None:
                    continue
                if not _same_occurrence(left, right) or not _overlap(left, right):
                    continue
                if _same_session(left, right):
                    continue
                pair_key = (
                    left.source["file"],
                    left.source["row"],
                    right.source["file"],
                    right.source["row"],
                    normalized_value,
                )
                if pair_key in seen_pairs:
                    continue
                seen_pairs.add(pair_key)
                collector.add(
                    "error",
                    code,
                    f"{label} {normalized_value!r} is assigned to overlapping sessions.",
                    [left.evidence(), right.evidence()],
                    shared_weeks=sorted(left.weeks.intersection(right.weeks)),
                )
    return usable_sessions, collector.counts.get(code, 0) - issue_count_before


def _check_time_quality(
    sessions: list[Session], collector: IssueCollector
) -> int:
    checked = 0
    for session in sessions:
        if session.start is None or session.end is None:
            collector.add(
                "error",
                "MISSING_OR_INVALID_TIME",
                "A scheduled record does not contain a usable start and end time.",
                [session.evidence()],
            )
            continue
        checked += 1
        if session.end <= session.start:
            collector.add(
                "error",
                "INVALID_TIME_RANGE",
                "A session ends at or before its start time.",
                [session.evidence()],
            )
    return checked


def _check_capacity(
    sessions: list[Session],
    room_inventory: dict[str, dict[str, Any]],
    collector: IssueCollector,
) -> int:
    checked = 0
    for session in sessions:
        if session.expected_students is None or not session.rooms:
            continue

        known_rooms = [
            room_inventory.get(room.casefold()) for room in session.rooms
        ]
        if room_inventory:
            for room_name, room in zip(session.rooms, known_rooms):
                if room is None:
                    collector.add(
                        "error",
                        "UNKNOWN_ROOM",
                        f"Room {room_name!r} is not present in the supplied room inventory.",
                        [session.evidence()],
                    )

        capacities = [
            room["capacity"]
            for room in known_rooms
            if room is not None and room["capacity"] is not None
        ]
        capacity = sum(capacities) if capacities else session.inline_room_capacity
        if capacity is None:
            continue
        checked += 1
        if session.expected_students > capacity:
            collector.add(
                "error",
                "ROOM_CAPACITY_EXCEEDED",
                "The expected student count exceeds the assigned room capacity.",
                [session.evidence()],
                expected_students=session.expected_students,
                available_capacity=capacity,
            )
    return checked


def _cohort_major_code(group_id: str) -> str | None:
    match = re.match(r"^(.+)-Y\d+-T\d+$", group_id.strip(), re.IGNORECASE)
    return match.group(1) if match else None


def _check_support_session_sharing(
    sessions: list[Session],
    collector: IssueCollector,
) -> int:
    checked = 0
    for session in sessions:
        normalized_type = _normalize_label(session.session_type)
        if normalized_type not in {"tutorial", "lab", "laboratory"}:
            continue
        checked += 1

        inferred_codes = {
            code.casefold()
            for group in session.student_groups
            if (code := _cohort_major_code(group))
        }
        listed_codes = {
            code.casefold() for code in session.major_codes if code
        }
        listed_majors = {
            major.casefold() for major in session.majors if major
        }

        violations: list[str] = []
        if len(session.student_groups) != 1:
            violations.append(
                "Tutorials and labs must contain exactly one cohort group."
            )
        if len(inferred_codes) > 1:
            violations.append(
                "The cohort groups belong to more than one major."
            )
        if len(listed_codes) > 1 or len(listed_majors) > 1:
            violations.append(
                "The row lists more than one major."
            )
        if (
            len(inferred_codes) == 1
            and len(listed_codes) == 1
            and inferred_codes != listed_codes
        ):
            violations.append(
                "The listed major code does not match the cohort group."
            )
        if (
            not session.student_groups
            or (not inferred_codes and not listed_codes and not listed_majors)
        ):
            violations.append(
                "The row does not contain enough major/cohort information to "
                "prove that it is major-specific."
            )

        if violations:
            collector.add(
                "error",
                "CROSS_MAJOR_SUPPORT_SESSION",
                (
                    "A tutorial or lab violates the rule that only lectures "
                    "may be shared across majors."
                ),
                [session.evidence()],
                violations=violations,
            )
    return checked


def _check_exam_rules(
    sessions: list[Session],
    rules: dict[str, Any],
    collector: IssueCollector,
) -> tuple[int, int]:
    finals_checked = 0
    quizzes_checked = 0
    allowed_final_times = {
        _time_minutes(value) for value in rules["final_exam_start_times"]
    }
    preferred_quiz_periods = {
        _period_key(value) for value in rules["preferred_quiz_periods"]
    }

    for session in sessions:
        normalized_type = _normalize_label(session.session_type)
        if "final" in normalized_type:
            finals_checked += 1
            if session.start not in allowed_final_times:
                collector.add(
                    "error",
                    "INVALID_FINAL_EXAM_TIME",
                    "A final exam does not start at an allowed final-exam time.",
                    [session.evidence()],
                    allowed_start_times=rules["final_exam_start_times"],
                )
        if "quiz" in normalized_type:
            quizzes_checked += 1
            if session.period not in preferred_quiz_periods:
                severity = (
                    "error" if rules["quiz_period_rule_is_hard"] else "warning"
                )
                collector.add(
                    severity,
                    "NON_PREFERRED_QUIZ_PERIOD",
                    "A quiz is outside the configured preferred periods.",
                    [session.evidence()],
                    preferred_periods=rules["preferred_quiz_periods"],
                )
    return finals_checked, quizzes_checked


def _same_slot(left: Session, right: Session) -> bool:
    return bool(
        _same_occurrence(left, right)
        and left.start is not None
        and left.end is not None
        and left.start == right.start
        and left.end == right.end
    )


def _room_booking_matches(
    session: Session,
    booking: Session,
    room_name: str,
) -> bool:
    return bool(
        session.session_id
        and booking.session_id
        and session.session_id.casefold() == booking.session_id.casefold()
        and room_name.casefold()
        in {value.casefold() for value in booking.rooms}
        and _same_slot(session, booking)
    )


def _check_room_schedule_consistency(
    sessions: list[Session],
    bookings: list[Session],
    collector: IssueCollector,
) -> int:
    checked = 0
    bookings_by_id: dict[str, list[Session]] = defaultdict(list)
    sessions_by_id: dict[str, list[Session]] = defaultdict(list)
    for booking in bookings:
        if booking.session_id:
            bookings_by_id[booking.session_id.casefold()].append(booking)
    for session in sessions:
        if session.session_id:
            sessions_by_id[session.session_id.casefold()].append(session)

    for session in sessions:
        if not session.session_id or not session.rooms:
            continue
        checked += 1
        candidates = bookings_by_id.get(session.session_id.casefold(), [])
        for room_name in session.rooms:
            if any(
                _room_booking_matches(session, booking, room_name)
                for booking in candidates
            ):
                continue
            collector.add(
                "error",
                "ROOM_SCHEDULE_MISMATCH",
                (
                    "A timetable session does not have a matching room "
                    "reservation with the same ID, room, day/week, and time."
                ),
                [session.evidence()],
                expected_room=room_name,
            )

    for booking in bookings:
        if not booking.session_id or not booking.rooms:
            continue
        candidates = sessions_by_id.get(booking.session_id.casefold(), [])
        for room_name in booking.rooms:
            if any(
                _room_booking_matches(session, booking, room_name)
                for session in candidates
            ):
                continue
            collector.add(
                "error",
                "ORPHAN_ROOM_RESERVATION",
                (
                    "A room reservation does not have a matching timetable "
                    "session with the same ID, room, day/week, and time."
                ),
                [booking.evidence()],
                reserved_room=room_name,
            )
    return checked


def _doctor_entry_matches(general: Session, doctor: Session) -> bool:
    general_staff = {value.casefold() for value in general.instructors}
    doctor_staff = {value.casefold() for value in doctor.instructors}
    general_course = (general.course_id or general.course_name or "").casefold()
    doctor_course = (doctor.course_id or doctor.course_name or "").casefold()
    general_rooms = {value.casefold() for value in general.rooms}
    doctor_rooms = {value.casefold() for value in doctor.rooms}
    return bool(
        general_staff.intersection(doctor_staff)
        and general_course
        and general_course == doctor_course
        and (not general_rooms or not doctor_rooms or general_rooms.intersection(doctor_rooms))
        and _same_slot(general, doctor)
    )


def _check_doctor_schedule_consistency(
    sessions: list[Session],
    doctor_sessions: list[Session],
    collector: IssueCollector,
) -> int:
    known_doctors = {
        instructor.casefold()
        for session in doctor_sessions
        for instructor in session.instructors
    }
    general_with_staff = [
        session
        for session in sessions
        if any(
            instructor.casefold() in known_doctors
            for instructor in session.instructors
        )
    ]
    checked = len(general_with_staff)
    doctors_by_name: dict[str, list[Session]] = defaultdict(list)
    general_by_name: dict[str, list[Session]] = defaultdict(list)
    for session in doctor_sessions:
        for instructor in session.instructors:
            doctors_by_name[instructor.casefold()].append(session)
    for session in general_with_staff:
        for instructor in session.instructors:
            general_by_name[instructor.casefold()].append(session)

    for session in general_with_staff:
        candidates: list[Session] = []
        for instructor in session.instructors:
            candidates.extend(doctors_by_name.get(instructor.casefold(), []))
        if any(_doctor_entry_matches(session, doctor) for doctor in candidates):
            continue
        collector.add(
            "error",
            "DOCTOR_SCHEDULE_MISMATCH",
            (
                "A staffed timetable session is missing from the matching "
                "doctor schedule or has different course, room, or time data."
            ),
            [session.evidence()],
        )

    for doctor in doctor_sessions:
        candidates = []
        for instructor in doctor.instructors:
            candidates.extend(general_by_name.get(instructor.casefold(), []))
        if any(_doctor_entry_matches(general, doctor) for general in candidates):
            continue
        collector.add(
            "error",
            "ORPHAN_DOCTOR_SESSION",
            (
                "A doctor schedule entry is missing from the general timetable "
                "or has different course, room, or time data."
            ),
            [doctor.evidence()],
        )
    return checked


def _check_result(
    name: str,
    checked_records: int,
    issue_count: int,
    unavailable_reason: str | None = None,
    warning_count: int = 0,
) -> dict[str, Any]:
    if unavailable_reason:
        status = "not_run"
    elif issue_count:
        status = "failed"
    elif warning_count:
        status = "warning"
    else:
        status = "passed"
    result: dict[str, Any] = {
        "name": name,
        "status": status,
        "records_checked": checked_records,
        "issue_count": issue_count + warning_count,
        "error_count": issue_count,
        "warning_count": warning_count,
    }
    if unavailable_reason:
        result["reason"] = unavailable_reason
    return result


def _mapping_request(region: Region, missing: list[str]) -> dict[str, Any]:
    is_room_source = region.role == "rooms"
    is_availability_source = region.role == "room_availability"
    is_staff_source = region.role == "staff_directory"
    example_mapping = (
        {
            "_role": "rooms",
            "room_id": "<uploaded room header>",
            "room_capacity": "<uploaded capacity header>",
            "room_type": "<uploaded room-type header>",
        }
        if is_room_source
        else {
            "_role": "room_availability",
            "session_id": "<uploaded reservation/session header>",
            "room_id": "<uploaded room header>",
            "day": "<uploaded day header>",
            "start": "<uploaded start-time header>",
            "end": "<uploaded end-time header>",
            "status": "<uploaded availability-status header>",
        }
        if is_availability_source
        else {
            "_role": "staff_directory",
            "staff_id": "<uploaded staff-ID header>",
            "instructor": "<uploaded staff-name header>",
        }
        if is_staff_source
        else {
            "_role": "sessions",
            "course_id": "<uploaded course header>",
            "day": "<uploaded day header>",
            "start": "<uploaded start-time header>",
            "end": "<uploaded end-time header>",
            "room_id": "<uploaded room header>",
            "instructor": "<uploaded instructor header>",
            "student_groups": "<uploaded group header>",
        }
    )
    return {
        "source_key": region.key,
        "reason": "Required validation fields could not be identified safely.",
        "missing_canonical_fields": missing,
        "available_headers": region.headers,
        "detected_mapping": region.mapping,
        "how_to_retry": {
            "column_mappings": {region.key: example_mapping}
        },
    }


def _compact_mapping_requests(
    requests: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    grouped: dict[tuple[Any, ...], list[dict[str, Any]]] = defaultdict(list)
    for request in requests:
        file_name = request["source_key"].split("::", 1)[0]
        signature = (
            file_name,
            tuple(request["missing_canonical_fields"]),
            tuple(request["available_headers"]),
            json.dumps(request["detected_mapping"], sort_keys=True),
        )
        grouped[signature].append(request)

    compact: list[dict[str, Any]] = []
    for (file_name, _, _, _), values in grouped.items():
        first = dict(values[0])
        first["matching_source_count"] = len(values)
        first["example_source_keys"] = [
            value["source_key"] for value in values[:10]
        ]
        if len(values) > 1:
            example_mapping = next(
                iter(first["how_to_retry"]["column_mappings"].values())
            )
            first["how_to_retry"] = {
                "mapping_scope": (
                    "The repeated layout can be mapped once at file scope."
                ),
                "column_mappings": {file_name: example_mapping},
            }
        compact.append(first)
    return compact


def _compact_source_profiles(regions: list[Region]) -> list[dict[str, Any]]:
    grouped: dict[tuple[Any, ...], list[Region]] = defaultdict(list)
    for region in regions:
        signature = (
            region.file_name,
            region.role,
            region.role_reason,
            tuple(region.headers),
            json.dumps(region.mapping, sort_keys=True),
        )
        grouped[signature].append(region)

    profiles: list[dict[str, Any]] = []
    for _, values in grouped.items():
        first = values[0]
        profiles.append(
            {
                "file": first.file_name,
                "role": first.role,
                "role_reason": first.role_reason,
                "matching_source_count": len(values),
                "total_rows": sum(len(region.rows) for region in values),
                "example_source_keys": [region.key for region in values[:10]],
                "headers": first.headers,
                "mapping": first.mapping,
            }
        )
    return profiles


def _rule_errors(rules: dict[str, Any]) -> list[str]:
    errors: list[str] = []
    teaching_weeks = rules.get("teaching_weeks")
    if (
        not isinstance(teaching_weeks, (list, tuple, set))
        or not teaching_weeks
        or any(_as_int(value) is None for value in teaching_weeks)
    ):
        errors.append("teaching_weeks must be a non-empty list of week numbers.")

    final_times = rules.get("final_exam_start_times")
    if (
        not isinstance(final_times, (list, tuple, set))
        or not final_times
        or any(_time_minutes(value) is None for value in final_times)
    ):
        errors.append(
            "final_exam_start_times must be a non-empty list of clock times."
        )

    quiz_periods = rules.get("preferred_quiz_periods")
    if not isinstance(quiz_periods, (list, tuple, set)) or not quiz_periods:
        errors.append(
            "preferred_quiz_periods must be a non-empty list of period labels."
        )
    if not isinstance(rules.get("quiz_period_rule_is_hard"), bool):
        errors.append("quiz_period_rule_is_hard must be true or false.")
    return errors


@tool
def check_validity(
    schedule_files: list[str],
    column_mappings: dict[str, dict[str, str]] | None = None,
    rules: dict[str, Any] | None = None,
    max_issues: int = 100,
) -> str:
    """Normalize user schedule files and exhaustively check them for conflicts.

    Call once with the relevant file names. The result contains source profiles
    and ``mapping_requests`` when unfamiliar headers cannot be understood
    safely. In that case, inspect the source with ``get_schedule``, ask the user
    to confirm uncertain meanings, and call this tool again with
    ``column_mappings``. Mapping keys may be a file name, ``file::sheet``, or the
    exact ``source_key`` returned by this tool. Each mapping maps canonical
    fields such as ``room_id`` or ``student_groups`` to uploaded headers and may
    include ``_role`` set to sessions, doctor_sessions, rooms,
    room_availability, periods, staff_directory, or ignore.
    """
    if not schedule_files:
        return _error(
            "no_schedule_files",
            "At least one schedule file must be supplied for validation.",
        )
    if not FAKE_DATA_DIR.is_dir():
        return _error(
            "fake_data_folder_missing",
            "The fake data folder does not exist.",
            expected_path=str(FAKE_DATA_DIR),
        )

    column_mappings = column_mappings or {}
    effective_rules = dict(DEFAULT_RULES)
    if rules:
        effective_rules.update(rules)
    invalid_rules = _rule_errors(effective_rules)
    if invalid_rules:
        return _error(
            "invalid_rules",
            "One or more validation rules have an invalid format.",
            problems=invalid_rules,
        )
    effective_rules["teaching_weeks"] = [
        _as_int(value) for value in effective_rules["teaching_weeks"]
    ]
    effective_rules["final_exam_start_times"] = [
        _format_minutes(_time_minutes(value))
        for value in effective_rules["final_exam_start_times"]
    ]
    effective_rules["preferred_quiz_periods"] = [
        str(value) for value in effective_rules["preferred_quiz_periods"]
    ]
    max_issues = max(1, min(int(max_issues), 1_000))

    resolved_paths: list[Path] = []
    source_errors: list[dict[str, Any]] = []
    unsupported_sources: list[dict[str, Any]] = []
    seen_paths: set[Path] = set()
    for file_name in schedule_files:
        path, resolution_error = _resolve_file(file_name)
        if resolution_error:
            source_errors.append(resolution_error)
            continue
        assert path is not None
        if path in seen_paths:
            continue
        seen_paths.add(path)
        if path.suffix.casefold() not in SUPPORTED_EXCEL_EXTENSIONS:
            unsupported_sources.append(
                {
                    "file": path.name,
                    "reason": (
                        "Full deterministic validation currently requires a "
                        "structured .xlsx or .xlsm workbook. PDF table layout "
                        "must first be converted or confirmed as structured data."
                    ),
                }
            )
            continue
        resolved_paths.append(path)

    if not resolved_paths:
        return _json(
            {
                "status": "ok",
                "validation_status": "inconclusive",
                "validation_complete": False,
                "source_errors": source_errors,
                "unsupported_sources": unsupported_sources,
                "required_action": (
                    "Provide at least one readable Excel schedule. Convert or "
                    "confirm PDF tables before full validation."
                ),
            }
        )

    regions: list[Region] = []
    extraction_errors: list[dict[str, Any]] = []
    for path in resolved_paths:
        try:
            regions.extend(_discover_regions(path, column_mappings))
        except PermissionError:
            extraction_errors.append(
                {
                    "file": path.name,
                    "code": "file_locked",
                    "message": "Close the workbook and retry validation.",
                }
            )
        except Exception as exc:
            extraction_errors.append(
                {
                    "file": path.name,
                    "code": "workbook_extraction_failed",
                    "exception_type": type(exc).__name__,
                    "message": str(exc),
                }
            )

    mapping_requests: list[dict[str, Any]] = []
    mapping_errors: list[dict[str, Any]] = []
    for region in regions:
        if region.mapping_problems:
            mapping_errors.append(
                {
                    "source_key": region.key,
                    "problems": region.mapping_problems,
                }
            )
        if region.role in {
            "sessions",
            "doctor_sessions",
            "rooms",
            "room_availability",
            "staff_directory",
        }:
            if region.role == "rooms":
                missing = _room_missing_mapping(region)
            elif region.role == "room_availability":
                missing = _availability_missing_mapping(region)
            elif region.role == "staff_directory":
                missing = _staff_directory_missing_mapping(region)
            else:
                missing = _session_missing_mapping(region)
            if missing:
                mapping_requests.append(_mapping_request(region, missing))

    usable_regions = [
        region
        for region in regions
        if region.role in {"sessions", "doctor_sessions"}
        and not _session_missing_mapping(region)
    ]
    period_map = _build_period_map(regions)
    staff_directory = _build_staff_directory(regions)
    sessions: list[Session] = []
    doctor_sessions: list[Session] = []
    for region in usable_regions:
        target = doctor_sessions if region.role == "doctor_sessions" else sessions
        for row_number, row in region.rows:
            target.append(
                _normalize_session(
                    region,
                    row_number,
                    row,
                    period_map,
                    staff_directory,
                    effective_rules,
                )
            )

    availability_sessions: list[Session] = []
    for region in regions:
        if (
            region.role != "room_availability"
            or _availability_missing_mapping(region)
        ):
            continue
        for row_number, row in region.rows:
            status = _normalize_label(
                _mapped_value(row, region.mapping, "status")
            )
            session_id = _mapped_value(row, region.mapping, "session_id")
            course = (
                _mapped_value(row, region.mapping, "course_id")
                or _mapped_value(row, region.mapping, "course_name")
            )
            if status in {"available", "free", "open"}:
                continue
            if session_id in (None, "") and course in (None, ""):
                continue
            availability_sessions.append(
                _normalize_session(
                    region,
                    row_number,
                    row,
                    period_map,
                    staff_directory,
                    effective_rules,
                )
            )

    # Identical session IDs in multiple calendar views represent the same
    # session and must not be reported as conflicts against themselves.
    unique_sessions: list[Session] = []
    seen_session_ids: set[str] = set()
    mirrored_session_count = 0
    for session in sessions:
        if session.session_id:
            normalized_id = session.session_id.casefold()
            if normalized_id in seen_session_ids:
                mirrored_session_count += 1
                continue
            seen_session_ids.add(normalized_id)
        unique_sessions.append(session)
    sessions = unique_sessions

    unique_doctor_sessions: list[Session] = []
    seen_doctor_sessions: set[tuple[Any, ...]] = set()
    for session in doctor_sessions:
        identity = (
            tuple(value.casefold() for value in session.instructors),
            (session.course_id or session.course_name or "").casefold(),
            (session.day or "").casefold(),
            session.date,
            tuple(sorted(session.weeks)),
            session.start,
            session.end,
            tuple(value.casefold() for value in session.rooms),
            tuple(value.casefold() for value in session.student_groups),
        )
        if identity in seen_doctor_sessions:
            mirrored_session_count += 1
            continue
        seen_doctor_sessions.add(identity)
        unique_doctor_sessions.append(session)
    doctor_sessions = unique_doctor_sessions

    room_inventory = _build_room_inventory(regions)
    collector = IssueCollector(max_issues=max_issues)
    checks: list[dict[str, Any]] = []

    time_before = collector.total
    time_checked = _check_time_quality(sessions + doctor_sessions, collector)
    checks.append(
        _check_result(
            "time_integrity",
            time_checked,
            collector.total - time_before,
            None if sessions or doctor_sessions else "No usable session table was found.",
        )
    )

    room_usable, room_issues = _check_grouped_overlaps(
        sessions,
        "rooms",
        "ROOM_DOUBLE_BOOKING",
        "Room",
        collector,
    )
    checks.append(
        _check_result(
            "room_double_bookings",
            room_usable,
            room_issues,
            None if room_usable else "No sessions with mapped room and time fields.",
        )
    )

    staff_usable, staff_issues = _check_grouped_overlaps(
        sessions,
        "instructors",
        "INSTRUCTOR_DOUBLE_BOOKING",
        "Instructor",
        collector,
    )
    doctor_usable, doctor_issues = _check_grouped_overlaps(
        doctor_sessions,
        "instructors",
        "INSTRUCTOR_DOUBLE_BOOKING",
        "Instructor",
        collector,
    )
    checks.append(
        _check_result(
            "instructor_double_bookings",
            staff_usable + doctor_usable,
            staff_issues + doctor_issues,
            (
                None
                if staff_usable or doctor_usable
                else "No sessions with mapped instructor and time fields."
            ),
        )
    )

    group_usable, group_issues = _check_grouped_overlaps(
        sessions,
        "student_groups",
        "STUDENT_GROUP_DOUBLE_BOOKING",
        "Student group",
        collector,
    )
    checks.append(
        _check_result(
            "student_group_double_bookings",
            group_usable,
            group_issues,
            (
                None
                if group_usable
                else "No sessions with mapped student-group and time fields."
            ),
        )
    )

    sharing_before = collector.total
    support_sessions_checked = _check_support_session_sharing(
        sessions, collector
    )
    checks.append(
        _check_result(
            "lecture_only_cross_major_sharing",
            support_sessions_checked,
            collector.total - sharing_before,
            (
                None
                if support_sessions_checked
                else "No mapped tutorial or lab sessions were found."
            ),
        )
    )

    room_consistency_before = collector.total
    room_consistency_checked = 0
    if availability_sessions:
        room_consistency_checked = _check_room_schedule_consistency(
            sessions, availability_sessions, collector
        )
    checks.append(
        _check_result(
            "room_schedule_consistency",
            room_consistency_checked,
            collector.total - room_consistency_before,
            (
                None
                if availability_sessions and room_consistency_checked
                else "No mapped room-reservation schedule was supplied."
            ),
        )
    )

    general_staff_names = {
        value.casefold()
        for session in sessions
        for value in session.instructors
    }
    doctor_staff_names = {
        value.casefold()
        for session in doctor_sessions
        for value in session.instructors
    }
    doctor_consistency_before = collector.total
    doctor_consistency_checked = 0
    doctor_identity_matchable = bool(
        general_staff_names.intersection(doctor_staff_names)
    )
    if doctor_sessions and doctor_identity_matchable:
        doctor_consistency_checked = _check_doctor_schedule_consistency(
            sessions, doctor_sessions, collector
        )
    checks.append(
        _check_result(
            "doctor_schedule_consistency",
            doctor_consistency_checked,
            collector.total - doctor_consistency_before,
            (
                None
                if doctor_sessions and doctor_identity_matchable
                else (
                    "No mapped doctor schedule was supplied, or its staff "
                    "identifiers could not be matched to the general timetable."
                )
            ),
        )
    )

    capacity_before = collector.total
    capacity_checked = _check_capacity(sessions, room_inventory, collector)
    checks.append(
        _check_result(
            "room_capacity",
            capacity_checked,
            collector.total - capacity_before,
            (
                None
                if capacity_checked
                else "Capacity, room, or expected-student data is unavailable."
            ),
        )
    )

    exam_error_before = collector.severity_counts.get("error", 0)
    exam_warning_before = collector.severity_counts.get("warning", 0)
    finals_checked, quizzes_checked = _check_exam_rules(
        sessions, effective_rules, collector
    )
    exam_error_count = (
        collector.severity_counts.get("error", 0) - exam_error_before
    )
    exam_warning_count = (
        collector.severity_counts.get("warning", 0) - exam_warning_before
    )
    checks.append(
        _check_result(
            "exam_and_quiz_times",
            finals_checked + quizzes_checked,
            exam_error_count,
            (
                None
                if finals_checked or quizzes_checked
                else "No mapped final-exam or quiz records were found."
            ),
            warning_count=exam_warning_count,
        )
    )

    incomplete_reasons = []
    if source_errors:
        incomplete_reasons.append("One or more requested files could not be found.")
    if extraction_errors:
        incomplete_reasons.append("One or more workbooks could not be extracted.")
    if unsupported_sources:
        incomplete_reasons.append("One or more PDF or unsupported sources were not validated.")
    if mapping_requests or mapping_errors:
        incomplete_reasons.append("One or more source mappings require confirmation.")
    if not sessions and not doctor_sessions:
        incomplete_reasons.append("No usable session records were found.")
    if any(check["status"] == "not_run" for check in checks):
        incomplete_reasons.append("One or more validation checks could not run.")

    validation_complete = not incomplete_reasons
    if collector.severity_counts.get("error", 0):
        validation_status = "invalid"
    elif not validation_complete:
        validation_status = "inconclusive"
    else:
        validation_status = "valid"

    source_profiles = _compact_source_profiles(regions)
    mapping_requests = _compact_mapping_requests(mapping_requests)

    return _json(
        {
            "status": "ok",
            "validation_status": validation_status,
            "validation_complete": validation_complete,
            "summary": {
                "requested_file_count": len(schedule_files),
                "workbooks_read": len(resolved_paths) - len(extraction_errors),
                "source_region_count": len(regions),
                "session_records_checked": len(sessions),
                "doctor_session_records_checked": len(doctor_sessions),
                "room_reservation_records_checked": len(availability_sessions),
                "staff_directory_records": len(staff_directory),
                "mirrored_sessions_ignored": mirrored_session_count,
                "room_inventory_records": len(room_inventory),
                "total_issues": collector.total,
                "returned_issues": len(collector.items),
                "issues_truncated": collector.total > len(collector.items),
                "issues_by_severity": dict(collector.severity_counts),
                "issues_by_code": dict(collector.counts),
            },
            "checks": checks,
            "issues": collector.items,
            "mapping_requests": mapping_requests,
            "mapping_errors": mapping_errors,
            "source_errors": source_errors,
            "extraction_errors": extraction_errors,
            "unsupported_sources": unsupported_sources,
            "incomplete_reasons": incomplete_reasons,
            "source_profiles": source_profiles,
            "rules_applied": effective_rules,
            "validation_scope": {
                "supported_checks": [
                    check["name"] for check in checks
                ],
                "not_currently_evaluated": [
                    "detailed laboratory-equipment compatibility",
                    "accessibility requirements",
                    "staff workload and scheduling fairness",
                    "institution-specific room-type suitability beyond capacity",
                ],
                "note": (
                    "A valid result applies only to supported checks that ran. "
                    "Do not claim that unevaluated policies were verified."
                ),
            },
            "required_action": (
                "Inspect the requested sources with get_schedule, confirm each "
                "uncertain column meaning with the user, then call check_validity "
                "again with column_mappings."
                if mapping_requests or mapping_errors
                else None
            ),
        }
    )
