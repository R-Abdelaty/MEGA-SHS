"""Excel schedule normalization, stable event IDs, and schedule hashing."""

from __future__ import annotations

import hashlib
import json
import os
import re
import threading
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict, Field

from mega_shs.api_models import EventStatus, EventType, ScheduleEventResponse
from mega_shs.errors import ApiContractError


PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_DATA_DIR = PROJECT_ROOT / "fake data"
DEFAULT_SEMESTER_START = date(2026, 7, 5)
AUTHORITATIVE_FILE_PATTERN = re.compile(r"^0[1-7]_.+\.xlsx$", re.IGNORECASE)
DAY_OFFSETS = {
    "sunday": 0,
    "monday": 1,
    "tuesday": 2,
    "wednesday": 3,
    "thursday": 4,
}


class InternalModel(BaseModel):
    model_config = ConfigDict(extra="forbid", use_enum_values=True)


class NormalizedEvent(InternalModel):
    public: ScheduleEventResponse
    source_event_id: str
    source_file: str
    source_sheet: str
    source_row: int


class NormalizedSchedule(InternalModel):
    source_version: str
    events: list[NormalizedEvent]
    rooms: list[str] = Field(default_factory=list)


def _canonical_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().casefold().split())


def stable_event_id(
    *,
    source_file: str,
    sheet_name: str,
    row_number: int,
    source_event_id: str,
    name: str,
    event_type: str,
    student_group: str,
    event_date: date,
    start_time: str,
    end_time: str,
    room: str,
) -> str:
    """Return a deterministic ID for one date-specific schedule event."""
    identity = "\x1f".join(
        [
            _canonical_text(source_file),
            _canonical_text(sheet_name),
            str(row_number),
            _canonical_text(source_event_id),
            _canonical_text(name),
            _canonical_text(event_type),
            _canonical_text(student_group),
            event_date.isoformat(),
            start_time,
            end_time,
            _canonical_text(room),
        ]
    )
    return f"event_{hashlib.sha256(identity.encode('utf-8')).hexdigest()[:16]}"


def authoritative_schedule_files(data_dir: Path = DEFAULT_DATA_DIR) -> list[Path]:
    if not data_dir.is_dir():
        raise ApiContractError(
            503,
            "SCHEDULE_NOT_LOADED",
            "The schedule data directory is not available.",
        )
    files = sorted(
        (
            path
            for path in data_dir.iterdir()
            if path.is_file()
            and not path.name.startswith("~$")
            and AUTHORITATIVE_FILE_PATTERN.match(path.name)
        ),
        key=lambda item: item.name.casefold(),
    )
    if not files:
        raise ApiContractError(
            503,
            "SCHEDULE_NOT_LOADED",
            "No authoritative schedule workbooks are available.",
        )
    return files


def schedule_source_hash(files: Iterable[Path]) -> str:
    """Hash canonical file names and exact workbook bytes in stable order."""
    digest = hashlib.sha256()
    resolved = sorted((Path(path) for path in files), key=lambda p: p.name.casefold())
    if not resolved:
        raise ValueError("At least one schedule file is required")
    for path in resolved:
        digest.update(path.name.casefold().encode("utf-8"))
        digest.update(b"\0")
        with path.open("rb") as handle:
            while chunk := handle.read(1024 * 1024):
                digest.update(chunk)
        digest.update(b"\0")
    return f"sha256:{digest.hexdigest()}"


def preview_schedule_hash(events: Iterable[ScheduleEventResponse]) -> str:
    rows = [
        event.model_dump(mode="json")
        for event in sorted(events, key=lambda item: item.id)
    ]
    payload = json.dumps(rows, sort_keys=True, separators=(",", ":"), ensure_ascii=False)
    return f"preview:sha256:{hashlib.sha256(payload.encode('utf-8')).hexdigest()}"


def _semester_start() -> date:
    raw = os.getenv("SCHEDULE_SEMESTER_START_DATE", "").strip()
    if not raw:
        return DEFAULT_SEMESTER_START
    try:
        value = date.fromisoformat(raw)
    except ValueError as exc:
        raise ApiContractError(
            503,
            "SCHEDULE_NOT_LOADED",
            "SCHEDULE_SEMESTER_START_DATE must be an ISO 8601 date.",
        ) from exc
    if value.weekday() != 6:
        raise ApiContractError(
            503,
            "SCHEDULE_NOT_LOADED",
            "SCHEDULE_SEMESTER_START_DATE must be a Sunday.",
        )
    return value


def _date_for(semester_start: date, week: int, day_name: str) -> date:
    offset = DAY_OFFSETS.get(_canonical_text(day_name))
    if offset is None:
        raise ValueError(f"Unsupported teaching day: {day_name!r}")
    return semester_start + timedelta(days=((week - 1) * 7) + offset)


def _time_text(value: Any) -> str:
    if isinstance(value, datetime):
        value = value.time()
    if isinstance(value, time):
        return value.strftime("%H:%M")
    text = str(value or "").strip()
    if re.fullmatch(r"\d{1,2}:\d{2}(?::\d{2})?", text):
        hour, minute, *_ = text.split(":")
        return f"{int(hour):02d}:{minute}"
    raise ValueError(f"Unsupported schedule time: {value!r}")


def _rows(path: Path, sheet_name: str, header_row: int = 4) -> Iterable[tuple[int, dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        headers = [cell.value for cell in sheet[header_row]]
        for row_number, values in enumerate(
            sheet.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            if not any(value not in (None, "") for value in values):
                continue
            yield row_number, {
                str(header).strip(): value
                for header, value in zip(headers, values)
                if header not in (None, "")
            }
    finally:
        workbook.close()


def _weeks(value: Any) -> list[int]:
    numbers = [int(item) for item in re.findall(r"\d+", str(value or ""))]
    if not numbers:
        return []
    if len(numbers) == 2 and "-" in str(value):
        return list(range(numbers[0], numbers[1] + 1))
    return numbers


def _general_type(value: Any) -> EventType:
    normalized = _canonical_text(value)
    if normalized == "lecture":
        return EventType.LECTURE
    # The existing calendar intentionally has four visual count categories.
    # Source labs are represented in the tutorial/support-session category.
    return EventType.TUTORIAL


def _exam_type(value: Any) -> EventType:
    return (
        EventType.EXAM
        if "final" in _canonical_text(value) or "exam" in _canonical_text(value)
        else EventType.QUIZ
    )


def _event(
    *,
    path: Path,
    sheet: str,
    row_number: int,
    source_event_id: str,
    name: str,
    event_type: EventType,
    room: str,
    student_group: str,
    event_date: date,
    start: str,
    end: str,
) -> NormalizedEvent:
    event_id = stable_event_id(
        source_file=path.name,
        sheet_name=sheet,
        row_number=row_number,
        source_event_id=source_event_id,
        name=name,
        event_type=event_type.value,
        student_group=student_group,
        event_date=event_date,
        start_time=start,
        end_time=end,
        room=room,
    )
    return NormalizedEvent(
        public=ScheduleEventResponse(
            id=event_id,
            name=name or source_event_id,
            room=room or "Unassigned",
            type=event_type,
            student_group=student_group or "Unspecified",
            date=event_date,
            start_time=start,
            end_time=end,
            status=EventStatus.ACTIVE,
        ),
        source_event_id=source_event_id,
        source_file=path.name,
        source_sheet=sheet,
        source_row=row_number,
    )


def _load_general(path: Path, semester_start: date) -> list[NormalizedEvent]:
    events: list[NormalizedEvent] = []
    for row_number, row in _rows(path, "Semester Timetable"):
        source_id = str(row.get("Session ID") or "").strip()
        if not source_id:
            continue
        start = _time_text(row.get("Start"))
        end = _time_text(row.get("End"))
        for week in _weeks(row.get("Weeks")):
            event_date = _date_for(semester_start, week, str(row.get("Day") or ""))
            events.append(
                _event(
                    path=path,
                    sheet="Semester Timetable",
                    row_number=row_number,
                    source_event_id=source_id,
                    name=str(row.get("Course Name") or source_id).strip(),
                    event_type=_general_type(row.get("Session Type")),
                    room=str(row.get("Room") or "").strip(),
                    student_group=str(row.get("Cohort Group(s)") or "").strip(),
                    event_date=event_date,
                    start=start,
                    end=end,
                )
            )
    return events


def _load_active_rooms(path: Path) -> list[str]:
    """Return authoritative active room names from the room inventory."""
    rooms: list[str] = []
    seen: set[str] = set()
    for _, row in _rows(path, "Room Inventory"):
        room = " ".join(str(row.get("Room") or "").split())
        status = _canonical_text(row.get("Status"))
        key = _canonical_text(room)
        if room and status not in {"inactive", "closed", "maintenance"} and key not in seen:
            rooms.append(room)
            seen.add(key)
    return rooms


def _load_exam_sheet(
    path: Path,
    semester_start: date,
    sheet: str,
    id_field: str,
    type_field: str,
) -> list[NormalizedEvent]:
    events: list[NormalizedEvent] = []
    for row_number, row in _rows(path, sheet):
        source_id = str(row.get(id_field) or "").strip()
        week = int(row.get("Week") or 0)
        if not source_id or week < 1:
            continue
        major = row.get("Major(s)") or row.get("Major") or ""
        year = row.get("Year")
        student_group = f"{major} · Year {year}" if year else str(major)
        events.append(
            _event(
                path=path,
                sheet=sheet,
                row_number=row_number,
                source_event_id=source_id,
                name=str(row.get("Course Name") or source_id).strip(),
                event_type=_exam_type(row.get(type_field)),
                room=str(row.get("Room Assignments (Students)") or "").strip(),
                student_group=student_group.strip(" ·"),
                event_date=_date_for(
                    semester_start, week, str(row.get("Day") or "")
                ),
                start=_time_text(row.get("Start")),
                end=_time_text(row.get("End")),
            )
        )
    return events


class ExcelScheduleLoader:
    """Read-only loader with content-hash keyed in-process caching."""

    def __init__(self, data_dir: Path = DEFAULT_DATA_DIR) -> None:
        self.data_dir = Path(data_dir)
        self._lock = threading.RLock()
        self._cached: NormalizedSchedule | None = None

    def current_source_version(self) -> str:
        return schedule_source_hash(authoritative_schedule_files(self.data_dir))

    def load(self) -> NormalizedSchedule:
        files = authoritative_schedule_files(self.data_dir)
        source_version = schedule_source_hash(files)
        with self._lock:
            if self._cached and self._cached.source_version == source_version:
                return self._cached.model_copy(deep=True)

        by_name = {path.name: path for path in files}
        general = by_name.get("05_General_Schedule.xlsx")
        exams = by_name.get("06_Exam_Schedule.xlsx")
        rooms_file = by_name.get("01_Room_Schedule.xlsx")
        if general is None or exams is None or rooms_file is None:
            raise ApiContractError(
                503,
                "SCHEDULE_NOT_LOADED",
                "The room, general, or exam schedule workbook is missing.",
            )

        semester_start = _semester_start()
        rooms = _load_active_rooms(rooms_file)
        if not rooms:
            raise ApiContractError(
                503,
                "SCHEDULE_NOT_LOADED",
                "The authoritative room inventory contains no active rooms.",
            )
        events = _load_general(general, semester_start)
        events.extend(
            _load_exam_sheet(
                exams,
                semester_start,
                "Regular Assessments",
                "Assessment ID",
                "Assessment Type",
            )
        )
        events.extend(
            _load_exam_sheet(
                exams,
                semester_start,
                "Final Exams",
                "Exam ID",
                "Exam Type",
            )
        )
        events.sort(
            key=lambda item: (
                item.public.date,
                item.public.start_time,
                item.public.end_time,
                item.public.id,
            )
        )
        schedule = NormalizedSchedule(
            source_version=source_version,
            events=events,
            rooms=rooms,
        )
        with self._lock:
            self._cached = schedule.model_copy(deep=True)
        return schedule
