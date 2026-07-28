"""Stable numeric UI option mappings and authoritative dropdown catalogs."""

from __future__ import annotations

from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parent
FAKE_DATA_DIR = PROJECT_ROOT / "fake data"

DAYS = {
    1: "Sunday",
    2: "Monday",
    3: "Tuesday",
    4: "Wednesday",
    5: "Thursday",
}

PROBLEM_TYPES = {
    1: ("Doctor/TA unavailable", "lecturer_or_ta_unavailable"),
    2: ("Room/Lab unavailable", "room_closed"),
    3: ("Equipment unavailable", "equipment_unavailable"),
    4: ("Full university day cancelled", "day_cancelled"),
    5: ("Partial day/time block cancelled", "partial_day_cancelled"),
    6: ("Specific session cancelled", "session_cancelled"),
    7: ("Visiting professor availability changed", "visiting_professor_limited"),
    8: ("Unexpected exam/quiz added", "unexpected_exam"),
    9: ("University event using resources", "university_event"),
    10: ("Room capacity corrected", "room_capacity_corrected"),
    11: ("Proposed repair rejected", "repair_rejected"),
}

TIME_SCOPES = {
    1: "one_period",
    2: "period_range",
    3: "full_day",
}

PERIODS = {
    1: {"period_id": "P1", "label": "1st Period", "start": "08:30", "end": "10:00"},
    2: {"period_id": "P2", "label": "2nd Period", "start": "10:15", "end": "11:45"},
    3: {"period_id": "P3", "label": "3rd Period", "start": "12:00", "end": "13:30"},
    4: {"period_id": "P4", "label": "4th Period", "start": "13:45", "end": "15:15"},
    5: {"period_id": "P5", "label": "5th Period", "start": "15:45", "end": "17:15"},
}

CANCELLATION_REASONS = {
    1: "University event/day off",
    2: "Public holiday",
    3: "Emergency closure",
    4: "Severe weather",
    5: "Facility maintenance",
    6: "Safety incident",
    7: "Other",
}

ASSESSMENT_TYPES = {1: "Exam", 2: "Quiz", 3: "Lab test"}
URGENCIES = {1: "critical", 2: "high", 3: "normal", 4: "low"}
CONFIRMATIONS = {1: "create_prototype", 2: "cancel_request"}

REQUIRED_FIELDS = {
    1: ["day_option", "academic_week", "resource_options", "scope_option"],
    2: ["day_option", "academic_week", "resource_options", "scope_option"],
    3: ["day_option", "academic_week", "resource_options", "scope_option"],
    4: ["day_option", "academic_week", "reason_option", "confirmation_option"],
    5: ["day_option", "academic_week", "scope_option", "reason_option"],
    6: ["session_options"],
    7: ["day_option", "academic_week", "resource_options", "scope_option"],
    8: ["day_option", "academic_week", "assessment_option", "student_group_options", "start_period_option"],
    9: ["day_option", "academic_week", "resource_values", "scope_option"],
    10: ["resource_options", "corrected_room_capacity"],
    11: ["related_repair_id", "description"],
}


def _option_rows(mapping: dict[int, Any]) -> list[dict[str, Any]]:
    rows = []
    for option, value in mapping.items():
        if isinstance(value, tuple):
            label, backend_value = value
            rows.append({"option": option, "label": label, "value": backend_value})
        elif isinstance(value, dict):
            rows.append({"option": option, **value})
        else:
            rows.append({"option": option, "label": str(value), "value": value})
    return rows


def ui_options() -> dict[str, Any]:
    return {
        "api_version": "v1",
        "days": _option_rows(DAYS),
        "academic_weeks": [
            {"option": week, "label": f"Week {week}", "value": week}
            for week in range(1, 13)
        ],
        "problem_types": [
            {**row, "required_fields": REQUIRED_FIELDS[row["option"]]}
            for row in _option_rows(PROBLEM_TYPES)
        ],
        "time_scopes": _option_rows(TIME_SCOPES),
        "periods": _option_rows(PERIODS),
        "cancellation_reasons": _option_rows(CANCELLATION_REASONS),
        "assessment_types": _option_rows(ASSESSMENT_TYPES),
        "urgencies": _option_rows(URGENCIES),
        "confirmations": _option_rows(CONFIRMATIONS),
    }


def _read_rows(file_name: str, sheet_name: str, header_row: int = 4) -> list[dict[str, Any]]:
    workbook = load_workbook(FAKE_DATA_DIR / file_name, read_only=False, data_only=True)
    try:
        sheet = workbook[sheet_name]
        headers = [cell.value for cell in sheet[header_row]]
        result: list[dict[str, Any]] = []
        for values in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            if not any(value not in (None, "") for value in values):
                continue
            result.append(
                {
                    str(header).strip(): value
                    for header, value in zip(headers, values)
                    if header not in (None, "")
                }
            )
        return result
    finally:
        workbook.close()


def _catalog_options(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [{"option": index, **item} for index, item in enumerate(items, start=1)]


@lru_cache(maxsize=8)
def catalog(name: str) -> list[dict[str, Any]]:
    normalized = name.strip().casefold()
    if normalized == "staff":
        doctors = _read_rows("07_Doctor_Schedule_Calendar.xlsx", "Doctor Directory")
        doctor_ids = {
            str(row.get("Doctor Name") or "").strip().casefold(): str(row.get("Doctor ID") or "").strip()
            for row in doctors
            if row.get("Doctor Name") not in (None, "")
        }
        schedule = _read_rows("05_General_Schedule.xlsx", "Semester Timetable")
        names = sorted(
            {str(row.get("Instructor") or "").strip() for row in schedule if row.get("Instructor") not in (None, "")},
            key=str.casefold,
        )
        return _catalog_options(
            [
                {
                    "value": name,
                    "label": (
                        f"{doctor_ids[name.casefold()]} — {name}"
                        if doctor_ids.get(name.casefold())
                        else name
                    ),
                    "staff_id": doctor_ids.get(name.casefold()),
                }
                for name in names
            ]
        )
    if normalized == "rooms":
        rows = _read_rows("01_Room_Schedule.xlsx", "Room Inventory")
        items = []
        for row in rows:
            if str(row.get("Status") or "").strip().casefold() != "active":
                continue
            room = str(row.get("Room") or "").strip()
            if not room:
                continue
            room_type = str(row.get("Type") or "").strip()
            capacity = row.get("Capacity")
            items.append(
                {
                    "value": room,
                    "label": f"{room} — {room_type} — {capacity} seats",
                    "room_type": room_type,
                    "capacity": capacity,
                }
            )
        return _catalog_options(sorted(items, key=lambda item: item["value"].casefold()))
    if normalized == "equipment":
        rows = _read_rows("02_Lab_Equipment.xlsx", "Equipment Inventory")
        items = []
        for row in rows:
            asset_id = str(row.get("Asset Group ID") or "").strip()
            if not asset_id:
                continue
            equipment = str(row.get("Equipment") or "").strip()
            lab = str(row.get("Lab") or "").strip()
            items.append(
                {
                    "value": asset_id,
                    "label": f"{asset_id} — {equipment} — {lab}",
                    "equipment": equipment,
                    "lab": lab,
                    "working_units": row.get("Working Units"),
                }
            )
        return _catalog_options(sorted(items, key=lambda item: item["value"].casefold()))
    if normalized == "student-groups":
        rows = _read_rows("03_Student_Enrollment.xlsx", "Cohort Groups")
        items = [
            {
                "value": str(row["Group ID"]).strip(),
                "label": f"{row['Group ID']} — {row.get('Major')} — Year {row.get('Year')}",
                "major": row.get("Major"),
                "year": row.get("Year"),
                "students": row.get("Students"),
            }
            for row in rows
            if row.get("Group ID") not in (None, "")
        ]
        return _catalog_options(sorted(items, key=lambda item: item["value"].casefold()))
    if normalized == "sessions":
        rows = _read_rows("05_General_Schedule.xlsx", "Semester Timetable")
        items = []
        for row in rows:
            session_id = str(row.get("Session ID") or "").strip()
            if not session_id:
                continue
            items.append(
                {
                    "value": session_id,
                    "label": (
                        f"{session_id} — {row.get('Course Name')} — "
                        f"{row.get('Day')} {row.get('Period')}"
                    ),
                    "day": row.get("Day"),
                    "period_id": row.get("Period ID"),
                    "session_type": row.get("Session Type"),
                    "course_id": row.get("Course ID"),
                }
            )
        return _catalog_options(sorted(items, key=lambda item: item["value"].casefold()))
    raise ValueError("catalog must be staff, rooms, equipment, sessions, or student-groups")


def resolve_catalog_options(name: str, option_ids: list[int] | None) -> list[str]:
    if not option_ids:
        return []
    lookup = {item["option"]: item["value"] for item in catalog(name)}
    missing = [option for option in option_ids if option not in lookup]
    if missing:
        raise ValueError(f"Unknown {name} option number(s): {missing}")
    return [lookup[option] for option in option_ids]


def resolve_day(option: int | None) -> str | None:
    if option is None:
        return None
    if option not in DAYS:
        raise ValueError("day_option must be between 1 and 5")
    return DAYS[option]


def resolve_period(option: int | None) -> dict[str, Any] | None:
    if option is None:
        return None
    if option not in PERIODS:
        raise ValueError("period option must be between 1 and 5")
    return PERIODS[option]
